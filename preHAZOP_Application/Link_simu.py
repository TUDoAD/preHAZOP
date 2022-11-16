# -*- coding: utf-8 -*-
"""
Created on Thu Dec  9 08:12:05 2021

@author: tim11
"""

def Link_simulation(Path_graph, Path_simulation, Path_graph_link_simu, Path_results):
    
    import xml.etree.ElementTree as ET
    import networkx as nx
    from openpyxl import load_workbook
    
    Results_workbook=load_workbook(Path_results)
    Error_log=Results_workbook['Error_log']        
    max_col=Error_log.max_column
    Header_index={}  
    for i in range(1, max_col + 1): 
        Header = Error_log.cell(row = 1, column = i).value
        Header_index[Header]=i
    for row in range(1, 100):        
        if Error_log.cell(row=row, column=Header_index['Warning/Error']).value==None:        
            Counter_error=row
            break
    
    mytree = ET.parse(Path_simulation)# load DEXPI-File
    myroot = mytree.getroot()
    graph=nx.read_graphml(Path_graph)
             
    #predefine attributes    
    Simu_Ty='n.a.'    
    Simu_MF='n.a.'
    Simu_unit_MF='n.a.'
    Simu_VF='n.a.'
    Simu_unit_VF='n.a.'
    Simu_T='n.a.'
    Simu_unit_T='n.a.'
    Simu_P='n.a.'
    Subs='n.a.'
    Simu_unit_P='n.a.'
    Sol_frac='n.a.'
    Liq_frac='n.a.'
    Vap_frac='n.a.'
    Simu_unit_density='n.a.'
    Simu_density='n.a.'
    
    #Create new attributes for every node
    for node in graph.nodes():
        graph.add_node(node, Type_simu=Simu_Ty, Part_simu='No')
        
    #Create new attributes for every edge    
    for edge in graph.edges():
        graph.add_edge(edge[0], edge[1], Part_simu='No', subs=Subs,
                       P_operation=Simu_P, P_operation_unit=Simu_unit_P, 
                       T_operation=Simu_T, T_operation_unit=Simu_unit_T,
                       Mass_flow=Simu_MF, Mass_flow_unit=Simu_unit_MF, 
                       Volume_flow=Simu_VF, Volume_flow_unit=Simu_unit_VF,
                       Density=Simu_density, Density_unit=Simu_unit_density,
                       sol_frac=Sol_frac, liq_frac=Liq_frac, vap_frac=Vap_frac)                         

                       
    #DATA EXTRACTION FROM RESULT REPORT
    
    edges_simu=[]
    nodes_simu=[]
    
    #parsing through all objects in simulation
    for Object in myroot.findall('Object'):
        N=Object.get('name')
               
        #predefine optional attributes
        Simu_Ty='n.a'
        Simu_MF='n.a.'
        Simu_unit_MF='n.a.'
        Simu_VF='n.a.'
        Simu_unit_VF='n.a.'
        Simu_T='n.a.'
        Simu_unit_T='n.a.'
        Simu_P='n.a.'
        Subs='n.a.'
        Simu_unit_P='n.a.'
        Sol_frac='n.a.'
        Liq_frac='n.a.'
        Vap_frac='n.a.' 
        Simu_unit_density='n.a.'
        Simu_density='n.a.'
        
        #Object is a Material streams
        if Object.get('type')=='Material Stream':            
            Substances=[]
            read_substance='no'
            
            #Streams in the simulation has to labeled in a certatain format 
            #(P&ID label from the source of stream, P&ID label from the destination of stream)
            #only if the format is given, the data of stream will be extracted
            if len(N.split(', '))>1:  
                FromID=N.split(', ')[0]
                ToID=N.split(', ')[1]  
                
                #Read properties
                for Object_property in Object.findall('Property'):#all properties 
                
                    if Object_property.get('name')=='Mass Flow':
                        Simu_MF=(Object_property.get('value').replace(',', '.'))
                        Simu_unit_MF=Object_property.get('units')
                        if Simu_MF=="":
                            Simu_MF='n.d.'
                            Simu_unit_MF='n.d.'
                    elif Object_property.get('name')=='Density (Mixture)':
                        Simu_density=(Object_property.get('value').replace(',', '.'))
                        Simu_unit_density=Object_property.get('units')
                        if Simu_density=="":
                            Simu_density='n.d.'
                            Simu_unit_density='n.d.'                            
                    elif Object_property.get('name')=='Volumetric Flow':
                        Simu_VF=(Object_property.get('value').replace(',', '.'))
                        Simu_unit_VF=Object_property.get('units')
                        if Simu_VF=="":
                            Simu_VF='n.d.'
                            Simu_unit_VF='n.d.'
                    elif Object_property.get('name')=='Temperature':
                        Simu_T=(Object_property.get('value').replace(',', '.'))
                        Simu_unit_T=Object_property.get('units')
                        if Simu_T=="":
                            Simu_T='n.d.'
                            Simu_unit_T='n.d.'
                    elif Object_property.get('name')=='Pressure':
                        Simu_P=(Object_property.get('value').replace(',', '.'))
                        Simu_unit_P=Object_property.get('units')
                        if Simu_P=="":
                            Simu_P='n.d.'
                            Simu_unit_P='n.d.'
                    #start property because after this property the substances follow        
                    elif Object_property.get('name')=='Molar Fraction (Vapor)' and Object_property.get('value')!="":
                        Vap_frac=(Object_property.get('value').replace(',', '.'))
                        if Vap_frac=="":
                            Vap_frac='n.d.'
                    #start property because after this property the substances follow        
                    elif Object_property.get('name')=='Molar Fraction (Overall Liquid)' and Object_property.get('value')!="":
                        Liq_frac=(Object_property.get('value').replace(',', '.'))
                        if Liq_frac=="":
                            Liq_frac='n.d.'
                    elif Vap_frac!='n.a.' and Liq_frac!='n.a.':#start property because after this property the substances follow
                        Sol_frac=str(1-float(Liq_frac)-float(Vap_frac))
                        if Sol_frac=="":
                            Sol_frac='n.d.'
                    #special way to extract the substances in stream    
                    if read_substance=='yes':
                        for Object_property_name in ['Density', 'Temperature', 'Pressure', 'Flow', 'Fraction', 'Molar', 
                                                     'Mixture', 'Vapor', 'Solid', 'Liquid']:
                            if Object_property_name in Object_property.get('name'):#end property because all Substances were read out
                                read_substance='no'#end signal
                                Subs=str(Substances)[1:-1].replace("'", "")#save the substances (saving the list is only possible as string)                       
                                break
                        try:
                            float(Object_property.get('value').replace(',','.'))
                            if float(Object_property.get('value').replace(',','.'))>0.01:
                                Substances.append(Object_property.get('name'))                                
                        except ValueError:
                            read_substance=='no'
                    #start property because after this property the substances follow            
                    elif Object_property.get('name')=='Mixture Molar Fraction' and Object_property.get('value')=="":
                        read_substance='yes'#start signal
                        
                #Data will be safed as attributes in the edge with same source/destination name                       
                paths=nx.all_simple_paths(graph, FromID, ToID)
                for path in paths:
                    Okay='Yes'
                    edges=[]
                    Paths_detail_level=[]
                    for n in range(0, len(path)-1):
                        step_FromID=path[n]
                        step_ToID=path[n+1]
                        edges.append((step_FromID, step_ToID))
                    
                    for edge in edges:
                        if graph.edges[edge]['group']=='Signal':
                            Okay='No'
                            break
                            
                    for node in path:
                        if graph.nodes[node]['sub_group_2']=='Safety valves' or graph.nodes[node]['Class']=='Rupture disk':
                            Okay='No'
                            break
                            
                    if Okay=='Yes':                        
                        for node in path[1:-1]:                        
                            if graph.nodes[node]['group'] in ['Pump', 'Heat exchanger', 'Vessel', 
                                                              'Column', 'Filter', '...']:
                                Okay='No'
                                Paths_detail_level.append(path)
                                break
                            
                    if Okay=='Yes':                            
                        for node in path:
                            graph.add_node(node, Part_simu='Yes')
                            
                        for edge in edges:
                            graph.add_edge(edge[0], edge[1], Part_simu='Yes', subs=Subs,
                                           P_operation=Simu_P, P_operation_unit=Simu_unit_P, 
                                           T_operation=Simu_T, T_operation_unit=Simu_unit_T,
                                           Mass_flow=Simu_MF, Mass_flow_unit=Simu_unit_MF, 
                                           Volume_flow=Simu_VF, Volume_flow_unit=Simu_unit_VF,
                                           Density=Simu_density, Density_unit=Simu_unit_density,
                                           sol_frac=Sol_frac, liq_frac=Liq_frac, vap_frac=Vap_frac)                         
            
        #Object is equipment
        else:            
            Simu_Ty=Object.get('type')
                        
            #Save the data as atttributes in the node with the same label       
            if N in graph.nodes():  
                graph.add_node(N, Type_simu=Simu_Ty, Part_simu='Yes')
                               
            else:
                nodes_simu.append(N)


    #WARNINGS
               
    if nodes_simu!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value=('At least one node in the simulation does not exist in the graph. Please check Spelling in simulation for the next run.')    
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_simu)
        Counter_error+=1        
    if edges_simu!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='At least one edge in the simulation does not exist in the graph. Please check Spelling in simulation for the next run.'      
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(edges_simu)
        Counter_error+=1
    if Paths_detail_level!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='At least one possible path was detected, which include an invalid node. Please check detail level of simulation.'      
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(Paths_detail_level)
        Counter_error+=1  


    #ASSIGN MAX/MIN TEMPERATURES AND PRESSURES TO THE NODES

    #Determine unit of T
    for edge in graph.edges():
        if graph.edges[edge]['T_operation'] not in ['n.a.', 'n.d.']:
            T_unit=graph.edges[edge]['T_operation_unit']
            break
        
    #Determine unit of P
    for edge in graph.edges():
        if graph.edges[edge]['T_operation'] not in ['n.a.', 'n.d.']:
            P_unit=graph.edges[edge]['P_operation_unit']
            break                  
    
    #Create new attributes (predefined) for ervery node 
    for node in graph.nodes():
        graph.add_node(node, T_max_operation='n.a.', T_max_operation_unit='n.a.' , T_min_operation='n.a.', 
                       T_min_operation_unit='n.a.', P_max_operation='n.a.', P_max_operation_unit='n.a.', 
                       P_min_operation='n.a.', P_min_operation_unit='n.a.')
    
    #Create a list of all edges surrounding the recent node
    for node in graph.nodes():
        Temperatures=[]
        Pressures=[]         
        edges=list(graph.in_edges(node))+list(graph.out_edges(node))
        
        #notice all the temperatures and pressures of the streams
        for edge in edges:
            if graph.edges[edge]['T_operation'] not in ['n.a.', 'n.d.'] and graph.edges[edge]['P_operation'] not in ['n.a.', 'n.d.']:
                if graph.edges[edge]['Class']!='Heat transfer medium':              
                    Temperatures.append(float(graph.edges[edge]['T_operation']))
                    Pressures.append(float(graph.edges[edge]['P_operation']))
                
        #Identificate and saving the min/max values (only if temperatures or pressures are given)
        if Temperatures!=[]:
            T_max=max(Temperatures)
            T_min=min(Temperatures)
            graph.add_node(node, T_max_operation=str(T_max), T_max_operation_unit=T_unit,
                           T_min_operation=str(T_min), T_min_operation_unit=T_unit)      
        if Pressures!=[]:
            P_max=max(Pressures)
            P_min=min(Pressures)
            graph.add_node(node, P_max_operation=str(P_max), P_max_operation_unit=P_unit, 
                           P_min_operation=str(P_min), P_min_operation_unit=P_unit)
 
            
    #ASSIGN DENSITY TO NODES          
    
    #Create new attributes (predefined) for ervery node 
    for node in graph.nodes():
        graph.add_node(node, Density='n.a.', Density_unit='n.a.')
    
    #Determine unit of density
    for edge in graph.edges():
        if graph.edges[edge]['Density'] not in ['n.a.', 'n.d.']:
            Density_unit=graph.edges[edge]['Density_unit']
            break    
        
    #Create a list of all edges surrounding the recent node
    for node in graph.nodes():
        Densities=[]
        edges=list(graph.in_edges(node))+list(graph.out_edges(node))
        
        #notice all densities of the streams
        for edge in edges:
            if graph.edges[edge]['Density'] not in ['n.a.', 'n.d.'] and graph.edges[edge]['sub_group']!='Heat transfer pipe':
                Densities.append(float(graph.edges[edge]['Density']))
                        
        #Identificate and saving the max value
        if Densities!=[]:
            Density=max(Densities)
            graph.add_node(node, Density=str(Density), Density_unit=Density_unit) 
 
            
    #ASSIGN FLOW TO NODES          
    
    #Create new attributes (predefined) for ervery node 
    for node in graph.nodes():
        graph.add_node(node, Mass_flow='n.a.', Mass_flow_unit='n.a.')
    
    #Determine unit of density
    for edge in graph.edges():
        if graph.edges[edge]['Mass_flow'] not in ['n.a.', 'n.d.']:
            Mass_flow_unit=graph.edges[edge]['Mass_flow_unit']
            break    
        
    #Create a list of all edges surrounding the recent node
    for node in graph.nodes():
        Mass_flows=[]
        edges=list(graph.in_edges(node))+list(graph.out_edges(node))
        
        #notice all densities of the streams
        for edge in edges:
            if graph.edges[edge]['Mass_flow'] not in ['n.a.', 'n.d.'] and graph.edges[edge]['sub_group']!='Heat transfer pipe':
                Mass_flows.append(float(graph.edges[edge]['Mass_flow']))
                        
        #Identificate and saving the max value
        if Mass_flows!=[]:
            Mass_flow=max(Mass_flows)
            graph.add_node(node, Mass_flow=str(Mass_flow), Mass_flow_unit=Mass_flow_unit) 

            
    #ASSIGN SUBSTANCES TO NODES             
    
    #Create new attributes (predefined) for ervery node 
    for node in graph.nodes():
        if graph.nodes[node]['Part_simu']=='Yes':
            graph.add_node(node, Subs_process='n.d.')
            graph.add_node(node, Sub_heat_transfer='n.d.')                
        else:
            graph.add_node(node, Sub_heat_transfer='n.a.')
            graph.add_node(node, Subs_process='n.a.')
    
    #Create a list of all edges surrounding the recent node
    for node in graph.nodes():
        if graph.nodes[node]['Part_simu']=='Yes':        
            Subs=[]
            Subs_heat_transfer='n.d.'
            edges=list(graph.in_edges(node))+list(graph.out_edges(node))
            
            #Distinguish between node with or without heat transfer medium           
            if graph.nodes[node]['group']=='Heat exchanger' or graph.nodes[node]['C_H_system'] in ['full-tube heating cooling coil', 'heating-cooling jacket', 'Semi-tube heating cooling coil']:
                if graph.nodes[node]['group']!='Electric Heater':
                    
                    #take over the subs of incoming or outgoing edge
                    for edge in edges:
                        if graph.edges[edge]['subs'] not in ['n.a.', 'n.d.']:
                            if graph.edges[edge]['sub_group']=='Heat transfer pipe':
                                Subs_heat_transfer=graph.edges[edge]['subs']#no list needed because medium is only one and does not change
                            else:    
                                Subs=Subs+graph.edges[edge]['subs'].split(', ') 
                                
                #extra case electric heater
                else:
                    Subs_heat_transfer='n.a.'
                    for edge in edges:
                        if graph.edges[edge]['subs'] not in ['n.a.', 'n.d.']:               
                            Subs=Subs+graph.edges[edge]['subs'].split(', ')
                            

            else:
                Subs_heat_transfer='n.a.'
                for edge in edges:
                    if graph.edges[edge]['subs'] not in ['n.a.', 'n.d.']:   
                        Subs=Subs+graph.edges[edge]['subs'].split(', ')
                     
                        
            graph.add_node(node, Sub_heat_transfer=Subs_heat_transfer)                                   
            #Entry preparation
            Subs=list(set(Subs)) 
            Subs=str(Subs)[1:-1].replace("'", "")        
            graph.add_node(node, Subs_process=Subs)  
                     

    Results_workbook.save(Path_results)            
    nx.write_graphml(graph, Path_graph_link_simu, encoding='utf-8', prettyprint=True, infer_numeric_types=False)