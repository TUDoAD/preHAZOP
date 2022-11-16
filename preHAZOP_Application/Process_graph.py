# -*- coding: utf-8 -*-
"""
Created on Tue Jan 11 12:02:46 2022

@author: tim11
"""

def Graph_smart(Path_graph, Path_Graph, Path_results):
    
    import networkx as nx
    from openpyxl import load_workbook
        
    Results_workbook=load_workbook(Path_results)
    Error_log=Results_workbook['Error_log']        
    max_col=Error_log.max_column
    Header_index={}  
    for i in range(1, max_col + 1): 
        Header = Error_log.cell(row = 1, column = i).value
        Header_index[Header]=i
    for row in range(1, 100):
        if Error_log.cell(row=row, column=Header_index['Warning/Error']).value==None:        
            Counter_error=row
            break
    
    Graph=nx.read_graphml(Path_graph)
    
    
    #OPERATION OF COLUMNS AND VESSELS
       
    dict_sub_group={'Column':{'Distillation column':['Schortcut Column', 'Distillation Column'],
                              'Ab/Desorption/Extraction':['Absorption Column']},
                    'Vessel':{'Reactor':['Conversion Reactor', 'Equilibrium Reactor', 'Gibbs Reactor', 
                                         'Plug-Flow Reactor (PFR)', 'Continuous stirred tank Reactor (CSTR)'],
                              'Storage tank':['Tank', 'Liquid Storage Tank']}}      
    
    #sorting into sub groups by the dict
    for node in Graph.nodes():
        grouped='No'
        for group in dict_sub_group:  
            if Graph.nodes[node]['group']==group:
                for sub_group in dict_sub_group[group]:
                    for Type in dict_sub_group[group][sub_group]:
                        if Type==Graph.nodes[node]['Type_simu']:                            
                            Graph.add_node(node, sub_group=sub_group)
                            grouped='Yes'
                            break
                if grouped=='Yes':
                    break
        if grouped=='Yes':
            break 
        
    for node in Graph.nodes(): 
        if Graph.nodes[node]['group']=='Vessel' and Graph.nodes[node]['Part_simu']=='No':
            Graph.nodes[node]['group']='Storage tank'


    #DISTINGUISH BETWEEN HEATING AND COOLING (only Heat exchanger)
         
    nodes_heat_exchanger=[]
                              
    for node in Graph.nodes():
        if Graph.nodes[node]['group'] in ['Heat exchanger'] and Graph.nodes[node]['sub_group_2']=='n.a.':
            Heat_cool='n.d.'            
            T_in='n.d.'
            T_out='n.d.'  
            
            #identificate the cooling/heating stream and notice in and outlet temperature
            Counter_in=0
            Counter_out=0
            for edge in Graph.in_edges(node):
                if Graph.edges[edge]['sub_group']!='Heat transfer pipe' and Graph.edges[edge]['Part_simu']=='Yes':
                    Counter_in+=1
                    T_in=float(Graph.edges[edge]['T_operation'])
                    Vap_frac_in=float(Graph.edges[edge]['vap_frac'])
            for edge in Graph.out_edges(node):
                if Graph.edges[edge]['sub_group']!='Heat transfer pipe' and Graph.edges[edge]['Part_simu']=='Yes':
                    Counter_out+=1
                    T_out=float(Graph.edges[edge]['T_operation']) 
                    Vap_frac_out=float(Graph.edges[edge]['vap_frac'])
                    
            #If everything worked, we can say if the node is cooling or heating by comparing the in and outlet temperatures       
            if T_in not in ['n.d.', 'n.a.'] and T_out not in ['n.d.', 'n.a.'] and Counter_in==1 and Counter_out==1:
                    
                    if float(T_out)>float(T_in):
                        Heat_cool='Heating'
                    elif float(T_out)<float(T_in):
                        Heat_cool='Cooling'
                    if float(Vap_frac_out)>float(Vap_frac_in):
                        Heat_cool='Evaporation'
                    elif float(Vap_frac_out)<float(Vap_frac_in):
                        Heat_cool='Condensation'
                        
            if Heat_cool=='n.d.':
                nodes_heat_exchanger.append(node)
            elif Heat_cool=='Heating':
                Graph.add_node(node, sub_group='Heat exchanger (heating)', sub_group_2='Heater')
            elif Heat_cool=='Cooling':
                Graph.add_node(node, sub_group='Heat exchanger (cooling)', sub_group_2='Cooler')                
            elif Heat_cool=='Evaporation':
                Graph.add_node(node, sub_group='Heat exchanger (heating)', sub_group_2='Evaporator') 
            elif Heat_cool=='Condensation':
                Graph.add_node(node, sub_group='Heat exchanger (cooling)', sub_group_2='Condenser')
        
    if nodes_heat_exchanger!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='At least one Heat exchanger could not be grouped by simulation data.'      
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_heat_exchanger)[1:-1].replace("'", "")
        Counter_error+=1        
        

    #UNIT CONVERSION EDGE
    
    Unit_dict={'kJ':[10**3, 'J'], 'kmol':[10**3, 'mol'], 'bar':[10**5, 'Pa'], 't':[10**3, 'kg'], 
               'kW':[10^3, 'W'], 'L':[10**-3, 'm3'], 'h':[3600, 's'], 'min':[60, 's'], 'cm2':[10**-4, 'm2'], 
               'cm':[10**-2, 'm'], 'mm':[10**-3, 'm']}
    
    for edge in Graph.edges():
        for Attribute_unit in ['T_operation_unit', 'P_operation_unit', 'Volume_flow_unit', 'Mass_flow_unit']:
            if Graph.edges[edge][Attribute_unit] not in ['n.d.', 'n.a.']:

                Attribute=Attribute_unit.replace('_'+Attribute_unit.split('_')[-1], '')
                
                Unit=Graph.edges[edge][Attribute_unit]
                Unit=Unit.replace('°', '')
                Unit=Unit.replace('.', '')
                
                if '/' in Unit:
                    Unit_split=Unit.split('/')
                    Unit_1=Unit_split[0]
                    Unit_2=Unit_split[1]
                       
                    Factor=1
                    for key in Unit_dict.keys():
                        if key in Unit_1:
                            Factor=Factor*Unit_dict[key][0]
                    for key in Unit_dict.keys():
                        if key in Unit_2:
                            Factor=Factor/Unit_dict[key][0]
                else:
                    Factor=1
                    for key in Unit_dict.keys():
                        if key in Unit:
                            Factor=Factor*Unit_dict[key][0]
                            
                for key in Unit_dict.keys():
                    if key in Unit:
                        Unit=Unit.replace(key, Unit_dict[key][1])
                        
                Graph.edges[edge][Attribute_unit]=Unit
                Graph.edges[edge][Attribute]=str((float(Graph.edges[edge][Attribute])*Factor))


    #UNIT CONVERSION NODES
    
    Unit_dict={'kJ':[10**3, 'J'], 'kmol':[10**3, 'mol'], 'bar':[10**5, 'Pa'], 't':[10**3, 'kg'], 
               'kW':[10^3, 'W'], 'L':[10**-3, 'm3'], 'h':[3600, 's'], 'min':[60, 's'], 
               'cm2':[10**-4, 'm2'], 'cm':[10**-2, 'm'], 'mm':[10**-3, 'm']}
    
    #all Unit attributes with a relevant entry
    for node in Graph.nodes():
        for Attribute_unit in ['T_min_design_unit', 'T_max_design_unit', 'P_min_design_unit', 
                               'P_max_design_unit', 'V_unit', 'T_min_operation_unit', 'T_max_operation_unit', 
                               'P_min_operation_unit', 'P_max_operation_unit', 'Mass_flow_unit']:
            
            if Graph.nodes[node][Attribute_unit] not in ['n.d.', 'n.a.']: 
                #Attribute name is the Attribute unit name without _unit at the end
                Attribute=Attribute_unit.replace('_'+Attribute_unit.split('_')[-1], '')               
                if Graph.nodes[node][Attribute] not in ['n.d.', 'n.a.']:
                    
                    Unit=Graph.nodes[node][Attribute_unit]
                    Unit=Unit.replace('°', '')
                    Unit=Unit.replace('.', '')
                    
                    #Consideration of the fraction line
                    if '/' in Unit:
                        Unit_split=Unit.split('/')
                        Unit_1=Unit_split[0]
                        Unit_2=Unit_split[1] 
                        
                        #Calculate Conversion Factor by dict 
                        Factor=1
                        for key in Unit_dict.keys():
                            if key in Unit_1:
                                Factor=Factor*Unit_dict[key][0]
                        for key in Unit_dict.keys():
                            if key in Unit_2:
                                Factor=Factor/Unit_dict[key][0]
                    else:
                        Factor=1
                        for key in Unit_dict.keys():
                            if key in Unit:
                                Factor=Factor*Unit_dict[key][0]
                    
                    #replace the non-SI with SI
                    for key in Unit_dict.keys():
                        if key in Unit:
                            Unit=Unit.replace(key, Unit_dict[key][1])
                    
                    #overwrite the attributes
                    Graph.nodes[node][Attribute_unit]=Unit
                    Graph.nodes[node][Attribute]=str(float(Graph.nodes[node][Attribute])*Factor)
 
     
    #TRANSLATE AND ASSIGN MSR  
    
    Letter_parameter={'TD':'Temperature difference', 'PD':'Pressure difference', 'FD':'Flow difference', 'LD':'Level difference',
                       'TF':'Temperature ratio', 'PF':'Pressure ratio', 'FF':'Flow ratio', 'LF':'Level ratio',
                       'TQ':'Temperature integer/sum', 'PQ':'Pressure integer/sum', 'FQ':'Flow integer/sum', 'LQ':'Level integer/sum',
                       'T':'Temperature', 'P':'Pressure', 'F':'Flow', 'L':'Level'}
    Letter_measurement={'I':'indication', 'R':'registration', 'A+':'High Alarm', 'A-':'Low Alarm', 'A+-':'High Alarm', 'A+-':'Low Alarm', 
                        'C':'Control', 'S+':'high Shut down', 'S-':'Low Shut down', 'S+-':'High Shut down', 'S+-':'Low Shut down'}

    #Create new attribute
    for node in Graph.nodes():
        if Graph.nodes[node]['Class']!='PCE Request':
            Graph.add_node(node, measurements='No')                
        else:
            Graph.add_node(node, measurements='n.a.')
            
    node_measurements=nx.get_node_attributes(Graph, 'measurements')
    
    #Select all MSR_nodes
    for node_1 in Graph.nodes():       
        if Graph.nodes[node_1]['Class']=='PCE Request' and Graph.nodes[node_1]['PID_label']:
            Counter=0
            Measurement_sources=[]
            Process_connections=[]
            Measurements=[] 
            Entries=[]            
            
            #Looking after Process connections within the edges of the MSR_node and count them
            edges=list(Graph.in_edges(node_1))+list(Graph.out_edges(node_1))
            for edge in edges:
                if Graph.edges[edge]['group']!='Signal':
                    Counter+=1
                    Process_connections.append(edge)            
            for Process_connection in Process_connections:
                for node_2 in Process_connection:
                    if Graph.nodes[node_2]['Class']!='PCE Request':
                        Measurement_sources.append(node_2)
                        
            if Counter>0:
                
                #Translate MSR with the dict above 
                for Letter in Letter_parameter:    
                    if Graph.nodes[node_1]['request'][0:2]==Letter:#With additional letter
                        Parameter=Letter_parameter[Letter]
                        break
                    elif Graph.nodes[node_1]['request'][0]==Letter:#without additional letter
                        Parameter=Letter_parameter[Letter]
                        break
                for Letter in Letter_measurement:
                    if Letter in Graph.nodes[node_1]['request']:
                        Measurements.append(Letter_measurement[Letter])
                    
                for Measurement in Measurements:
                    Entries.append(Parameter+' '+Measurement)
                
                #MSR simply connected to process                                
                if Counter==1: 
                    Source=Measurement_sources[0]
                     
                    #Consider the potencial valve between MSR and the real Measurement source, so take the neighbor of the valve
                    if Graph.nodes[Source]['group'] in ['Valves'] and Graph.nodes[Source]['Sub_class'] in ['Manually adjusted', 'General'] and len(list(nx.all_neighbors(Graph, Source)))==2:
                        for node_2 in list(nx.all_neighbors(Graph, Source)):
                            if Graph.nodes[node_2]['Class']!='PCE Request':
                                Source=node_2
                    
                    if Graph.nodes[Source]['group'] in ['Pipe tee', 'Fittings']:
                        Paths_outlet=[]
                        Paths_inlet=[]                        
                        for node_3 in Graph.nodes():
                            if Graph.nodes[node_3]['group'] in ['Vessel', 'Column', 'Centrifuge', 'Pump', 'Filter', 'Heat exchanger', '...']:
                                
                                #Search after a connection between to a potencial outlet source
                                try:                                        
                                    Path_outlet=nx.shortest_path(Graph, node_3, Source)   
                                    
                                    #No Signal is allowed within the connection
                                    Okay_outlet='Yes'
                                    for node_4 in Path_outlet[1:-1]:
                                        if Graph.nodes[node_4]['group'] in ['Vessel', 'Column', 'Centrifuge', 'Pump', 'Filter', 'Heat exchanger', '...']:
                                            Okay_outlet='No'
                                            break
                                    if Okay_outlet=='Yes':                                        
                                        for n in range(0, len(Path_outlet)-1):
                                            step_FromID=Path_outlet[n]
                                            step_ToID=Path_outlet[n+1] 
                                            if Graph.edges[step_FromID, step_ToID]['group']=='Signal':
                                                Okay_outlet='No'
                                                break
                                            
                                except nx.NetworkXNoPath:
                                    Okay_outlet='No'
                                
                                #Save a correct Path                                        
                                if Okay_outlet=='Yes':                                        
                                    Paths_outlet.append(Path_outlet) 

                                #The same with inlet    
                                try:                                         
                                    Path_inlet=nx.shortest_path(Graph, Source, node_3)
                                    Okay_inlet='Yes'
                                    for node_4 in Path_inlet[1:-1]:
                                        if Graph.nodes[node_4]['group'] in ['Vessel', 'Column', 'Centrifuge', 'Pump', 'Filter', 'Heat exchanger', '...']:
                                            Okay_inlet='No'
                                            break
                                    if Okay_inlet=='Yes':   
                                        for n in range(0, len(Path_inlet)-1):
                                            step_FromID=Path_inlet[n]
                                            step_ToID=Path_inlet[n+1] 
                                            if Graph.edges[step_FromID, step_ToID]['group']=='Signal':
                                                Okay_inlet='No'
                                                break
                                except nx.NetworkXNoPath:
                                    Okay_inlet='No'                            
                                if Okay_inlet=='Yes':                                        
                                    Paths_inlet.append(Path_inlet) 
                        
                        #Identificate the outlet/inlet node
                        if Paths_outlet!=[]:
                            Entries_outlet=[]
                            if Entries!=[]:
                                for n in range(0, len(Entries)):
                                    Entries_outlet.append('Outlet '+Entries[n])                           
                                for Path in Paths_outlet:
                                    Source=Path[0]                                                
                                    if node_measurements[Source]=='No':
                                        node_measurements[Source]=Entries_outlet
                                    else:
                                        node_measurements[Source]=node_measurements[Source]+Entries_outlet                            
                        else:
                            print('No outlet assigning could be made') 
                              
                        if Paths_inlet!=[]:
                            Entries_inlet=[]
                            if Entries!=[]:
                                for n in range(0, len(Entries)):
                                    Entries_inlet.append('Inlet '+Entries[n])                            
                                for Path in Paths_inlet:
                                    Source=Path[-1]                                        
                                    if node_measurements[Source]=='No':
                                        node_measurements[Source]=Entries_inlet
                                    else:
                                        node_measurements[Source]=node_measurements[Source]+Entries_inlet 
                        else:
                            print('No inlet assigning could be made')
                    
                    #no inlet/outlet case
                    else:
                        #Entry in dict as preparation for write in attribute later
                        if Entries!=[]:                       
                            if node_measurements[Source]=='No':
                                node_measurements[Source]=Entries
                            else:
                                node_measurements[Source]=node_measurements[Source]+Entries
                        
                #MSR multiple connected to process                 
                elif Counter==2:
                    Source_1=Measurement_sources[0]
                    Source_2=Measurement_sources[1]

                    #Consider the potencial valve between MSR and the real Measurement source, so take the neighbor of the valve                    
                    if Graph.nodes[Source_1]['group'] in ['Valves'] and Graph.nodes[Source_1]['Sub_class'] in ['Manually adjusted', 'General'] and len(list(nx.all_neighbors(Graph, Source_1)))==2:
                        for node_2 in list(nx.all_neighbors(Graph, Source_1)):
                            if Graph.nodes[node_2]['Class']!='PCE Request':
                                Source_1=node_2     
                    if Graph.nodes[Source_2]['group'] in ['Valves'] and Graph.nodes[Source_2]['Sub_class'] in ['Manually adjusted', 'General'] and len(list(nx.all_neighbors(Graph, Source_2)))==2:
                        for node_2 in list(nx.all_neighbors(Graph, Source_2)):
                            if Graph.nodes[node_2]['Class']!='PCE Request':
                                Source_2=node_2  
                    
                    #Lokk after the path between the two sources (from 1 to 2 or other way)
                    try:
                        paths=list(nx.all_simple_paths(Graph, Source_1, Source_2))
                        for path in paths:
                            for node_2 in path:
                                if Graph.nodes[node_2]['group'] in ['Pump', 'Filter']:
                                    Source=node_2
                    except nx.NetworkXNoPath:
                        print('Source could not be found')
                    try:
                        paths=list(nx.all_simple_paths(Graph, Source_2, Source_1))
                        for path in paths:
                            for node_2 in path:
                                if Graph.nodes[node_2]['group'] in ['Pump', 'Filter']:
                                    Source=node_2
                    except nx.NetworkXNoPath:
                        print('Source could not be found')
                    
                    #Entry in dict
                    if Entries!=[]: 
                        if node_measurements[Source]=='No':
                            node_measurements[Source]=Entries
                        else:
                            node_measurements[Source]=node_measurements[Source]+Entries
                        
                else:
                    print('Unknown case') 

    #Make the real entry as attribute
    for node in node_measurements.keys():
        if node_measurements[node] not in ['n.a.', 'No']:#only for nodes with entries
            Final_entry=node_measurements[node]
            Final_entry=set(Final_entry)
            Final_entry=str(Final_entry)[1:-1].replace("'", "")#Modify the style of entry
            Graph.add_node(node, measurements=Final_entry)                                        

                        
    #TRANSLATE AND ASSIGN MSR (SIGNALS)
    
    Letter_parameter={'TD':'Temperature difference', 'PD':'Pressure difference', 'FD':'Flow difference', 'LD':'Level difference',
                       'TF':'Temperature ratio', 'PF':'Pressure ratio', 'FF':'Flow ratio', 'LF':'Level ratio',
                       'TQ':'Temperature integer/sum', 'PQ':'Pressure integer/sum', 'FQ':'Flow integer/sum', 'LQ':'Level integer/sum',
                       'T':'Temperature', 'P':'Pressure', 'F':'Flow', 'L':'Level'}
    Letter_signal={'C':'Control', 'S+':'Shut down (high)', 'S-':'Shut down (low)', 'S+-':'high Shut down', 'S+-':'low Shut down'}
    
    #Create new attribute
    for node in Graph.nodes():
        if Graph.nodes[node]['Class']!='PCE Request':            
            Graph.add_node(node, signals='No')                
        else:
            Graph.add_node(node, signals='n.a.')
            
    node_signals=nx.get_node_attributes(Graph, 'signals')    
    
    #Select the exact case of a simply connected MSR with one Signal to a non_MSR node
    #MSR_DIN=input('Which MSR DIN is used? ')
    for node_1 in Graph.nodes():        
        if Graph.nodes[node_1]['Class']=='PCE Request':            
            Counter_in=0
            Counter_out=0            
            Signals=[] 
            Entries_1=[]
            Entries_2=[]            
            
            #Looking after Signals within the edges of the MSR_node and count them (in and out)
            for edge in Graph.in_edges(node_1):
                if Graph.edges[edge]['group']=='Signal':
                    Counter_in+=1
            for edge in Graph.out_edges(node_1):
                if Graph.edges[edge]['group']=='Signal':
                    Counter_out+=1
                    Signal_edge=edge
            
            #This case is of interest
            if Counter_in==0 and Counter_out==1:
                
                #Translate MSR with the dict above
                for Letter in Letter_parameter:    
                    if Graph.nodes[node_1]['request'][0:2]==Letter:
                        Parameter=Letter_parameter[Letter]
                        break
                    elif Graph.nodes[node_1]['request'][0]==Letter:
                        Parameter=Letter_parameter[Letter]
                        break                        
                for Letter in Letter_signal:
                    if Letter in Graph.nodes[node_1]['request']:
                        Signals.append(Letter_signal[Letter])
                          
                for Signal in Signals:
                    Entries_1.append(Parameter+' '+Signal)

                #Identificate the destination node from the above identified outgoing signal
                for node_2 in Signal_edge:
                    if node_2!=node_1:
                        Signal_destination_1=node_2
                    
                #If an equipment follow, it is a simple control
                if Graph.nodes[Signal_destination_1]['Class']!='PCE Request':
                    
                    #Entry in dict as preparation for write in attribute later 
                    if Entries_1!=[]:                     
                        if node_signals[Signal_destination_1]=='No':
                            node_signals[Signal_destination_1]=Entries_1
                        else:
                            node_signals[Signal_destination_1]=node_signals[Signal_destination_1]+Entries_1    
                
                #otherwise it can be a Cascade
                else:
                    for edge in Graph.in_edges(Signal_destination_1):
                        if Graph.edges[edge]['group']=='Signal':
                            Counter_in+=1
                    for edge in Graph.out_edges(Signal_destination_1):
                        if Graph.edges[edge]['group']=='Signal':
                            Counter_out+=1
                            Signal_edge=edge
 
                    if Counter_in==1 and Counter_out==1:
                        
                        #Translate MSR with the dict above
                        for Letter in Letter_parameter:    
                            if Graph.nodes[Signal_destination_1]['request'][0-2]==Letter:
                                Parameter=Letter_parameter[Letter]
                                break
                            elif Graph.nodes[Signal_destination_1]['request'][0]==Letter:
                                Parameter=Letter_parameter[Letter]
                                break
                        for Letter in Letter_signal:
                            if Letter in Graph.nodes[Signal_destination_1]['request']:
                                Signals.append(Letter_signal[Letter])                            
                        for Signal in Signals:
                            Entries_2.append(Parameter+' '+Signal)
                        
                        for node_2 in Signal_edge:
                            if node_2!=node_1:
                                Signal_destination_2=node_2
                        
                        #If an equipment follows it is a cascade
                        if Graph.nodes[Signal_destination_2]['Class']!='PCE Request':
                            
                            #Entry in dict as preparation for write in attribute later
                            if Entries_1!=[] and Entries_2!=[]:
                                if node_signals[node_1]=='No':
                                    node_signals[node_1]=Entries_1+['by']+Entries_2
                                else:
                                    node_signals[node_1]=node_signals[node_1]+Entries_1+['by']+Entries_2
                        
                        #otherwise it is something unknwon
                        else:
                            print('Unknown case')
                        
    #Make the real entry as attribute
    for node in Graph.nodes():
        if node_signals[node] not in ['n.a.', 'No']:#only for nodes with entries
            Final_entry=node_signals[node]
            set(Final_entry)
            Final_entry=str(Final_entry)[1:-1].replace("'", "")#Modify the style of entry
            Final_entry=Final_entry.replace(", by, ", " by ")#Modify the style of entry
            Graph.add_node(node, signals=Final_entry)  
  
    
    #REMOVE SIGNAL LINES
    
    edges=[]
    
    for edge in Graph.edges():
        if Graph.edges[edge]['group']=='Signal':
            edges.append(edge)
            
    Graph.remove_edges_from(edges)
    
    
    #REMOVE MSR AND CONNECTED VALVES
    
    list_MSR=[]
    list_MSR_valve=[]
    
    for node in Graph.nodes():
        if 'PIF' in Graph.nodes[node]['DEXPI_ID']:
            list_MSR.append(node)
            all_neighbors=list(nx.all_neighbors(Graph, node)) 
            if len(all_neighbors)==1:
                node_neighbor=all_neighbors[0]
                all_neighbors=list(nx.all_neighbors(Graph, node_neighbor)) 
                if len(all_neighbors)==2 and Graph.nodes[node_neighbor]['group']=='Valves':
                    list_MSR_valve.append(node_neighbor)
                      
    Graph.remove_nodes_from(list_MSR)      
    Graph.remove_nodes_from(list_MSR_valve)
    
    
    #CREATE NEW ATTRIBUTE AND DICT FOR FOLLOWING SAFEGUARDS
    
    for node in Graph.nodes():
        if Graph.nodes[node]['Class']!='PCE Request':            
            Graph.add_node(node, safeguards='No')        
        else:
            Graph.add_node(node, safeguards='n.a.') 
                
    node_safeguards=nx.get_node_attributes(Graph, 'safeguards') 

           
    #BYPASS FOR CONTROL VALVE AND PUMP

    Paths_remove=[]
    Cycles=nx.cycle_basis(Graph.to_undirected())

    #search through Cycles regarding Bypasses
    for Cycle in Cycles:
        Counter_control_valve=0
        Counter_pump=0
        Counter_other=0
        
        for node in Cycle:
            if Graph.nodes[node]['group']=='Pump':
                Pump=node
                Counter_pump+=1                 
            if Graph.nodes[node]['sub_group']=='Valves (operation)' and Graph.nodes[node]['Sub_class'] not in ['Manually adjusted', 'General']:
                Control_valve=node
                Counter_control_valve+=1 
            elif Graph.nodes[node]['group'] not in ['Valves', 'Fittings', 'Column'] and Graph.nodes[node]['Class']!='Pipe tee':#no Vessel or something like that in the path
                Counter_other+=1
            
        if Counter_pump==1 and Counter_other==0:
            Subgraph=Graph.subgraph(Cycle)
            Counter_start_node=0
            Counter_end_node=0
            for node in Cycle:
                if len(Subgraph.in_edges(node))==0:
                    Start_node=node
                    Counter_start_node+=1
                elif len(Subgraph.out_edges(node))==0:
                    End_node=node
                    Counter_end_node+=1
                    
            if Counter_start_node==1 and Counter_end_node==1: 
                Paths=list(nx.all_simple_paths(Graph, Start_node, End_node))#all paths
                
                for Path in Paths:
                    for node in Path:
                        if node==Pump:
                            Path_keep=Path
                            
                for Path in Paths:
                    if Path!=Path_keep:
                        Paths_remove=Paths_remove+Path[1:-1]
                        
                if type(node_safeguards[Pump])==list:
                     node_safeguards[Pump].append('Bypass')                               
                else:
                    node_safeguards[Pump]=['Bypass']                  
                
            else:
                Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='The identification of start and end node of this cycle was not possible.'   
                Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(Cycle)
                Counter_error+=1

        #subgraph of found cycle is created to identificate start end end node of the bypass 
        elif Counter_control_valve==1 and Counter_other==0:
           Subgraph=Graph.subgraph(Cycle)
           Counter_start_node=0
           Counter_end_node=0
           for node in Cycle:
               if len(Subgraph.in_edges(node))==0:
                   Start_node=node
                   Counter_start_node+=1
               elif len(Subgraph.out_edges(node))==0:
                   End_node=node
                   Counter_end_node+=1
                   
        #When start and end is clear, the bypass can be removed
           if Counter_start_node==1 and Counter_end_node==1:
                Paths=list(nx.all_simple_paths(Graph, Start_node, End_node))#all paths
                                
                for Path in Paths:
                    for node in Path:
                        if node==Control_valve:
                            Path_keep=Path
                            break
                        
                for Path in Paths:
                    if Path!=Path_keep:
                        Paths_remove=Paths_remove+Path[1:-1]
                        
                if type(node_safeguards[Control_valve])==list:
                     node_safeguards[Control_valve].append('Bypass')                               
                else:
                    node_safeguards[Control_valve]=['Bypass']
                     
           else:
                Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='The identification of start and end node of this cycle was not possible.'   
                Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(Cycle)
                Counter_error+=1    
    
    Graph.remove_nodes_from(Paths_remove)    


    #REDUNTANT PUMPS
    
    Start_nodes=[]
    End_nodes=[]
    Nodes_remove=[]
        
    #Select possible start and end points of a reduntant pump
    for node in Graph.nodes():
        if Graph.nodes[node]['Class']=='Pipe tee' and len(Graph.in_edges(node))==1 and len(Graph.out_edges(node))>=2:
            Start_nodes.append(node)
        elif Graph.nodes[node]['group']=='Valves' and len(Graph.in_edges(node))==1 and len(Graph.out_edges(node))>=2:
            Start_nodes.append(node)
        elif Graph.nodes[node]['Class']=='Pipe tee' and len(Graph.in_edges(node))>=2 and len(Graph.out_edges(node))==1:
            End_nodes.append(node)
        elif Graph.nodes[node]['group']=='Valves' and len(Graph.in_edges(node))>=2 and len(Graph.out_edges(node))==1:
            End_nodes.append(node)            

    #select possible paths between these nodes        
    for Start_node in Start_nodes:
        for End_node in End_nodes:
            Paths=list(nx.all_simple_paths(Graph, Start_node, End_node))
            if len(Paths)>=2:
                
                #check the paths which belongs potentially part of reduntant pump
                Redundant_path=[]
                Save_nodes=[]
                for Path in Paths:
                    Counter=0
                    path_clean='Yes'                   
                    for Path_node in Path:
                        if Graph.nodes[Path_node]['group']=='Pump':#one pump has to be found in the path
                            Counter+=1
                        elif Graph.nodes[Path_node]['group'] not in ['Valves', 'Fittings'] and Graph.nodes[Path_node]['Class']!='Pipe tee':#no Vessel or something like that in the path
                            path_clean='No'
                        if Path_node in Save_nodes:#the paths are not allowed to have any common nodes
                            path_clean='No'
                            
                    #Check if one pump is within each path and all paths were correct        
                    if Counter==1 and path_clean=='Yes':
                        Redundant_path.append('Yes')
                    else:
                        Redundant_path.append('No')
                        
                    #save nodes within the path for later checking in the next path
                    Path_shortened=Path.copy()
                    Path_shortened.remove(Start_node)
                    Path_shortened.remove(End_node)
                    Save_nodes=Save_nodes+Path_shortened
                
                #remove the redundant pump and its path
                if 'No' not in Redundant_path:
                    Paths=list(nx.all_simple_paths(Graph, Start_node, End_node))#all paths
                    Path_keep=[]
                    for Path in Paths:                        
                        for node in Path:
                            if Graph.nodes[node]['group']=='Pump' and Graph.nodes[node]['Part_simu']=='Yes':
                                Path_keep=Path
                                break
                            
                    if Path_keep==[]:
                        print('Redundant path could not be found')                        
                    else:        
                        #Path_keep=max(Paths, key=len)#the path which we want to keep in graph is the longest one
                        Paths.remove(Path_keep)
                        
                        #Notice the redundance of the pump which remains in graph
                        for node in Path_keep:
                            if Graph.nodes[node]['group']=='Pump':
                                if type(node_safeguards[node])==list:
                                     node_safeguards[node].append('Redundant')                               
                                else:
                                    node_safeguards[node]=['Redundant']                           
                        
                        #One big list of all nodes out of all paths which should be removed
                        for Path_remove in Paths:
                            Nodes_remove=Nodes_remove+Path_remove[1:-1]                        
                                            
    Graph.remove_nodes_from(Nodes_remove)


    #ASSIGN SAFETY VALVE AND RUPTURE DISC
    
    Already_assigned=[]
    No_assign=[]
    
    for node in Graph.nodes():
        Rupture_disc='No'
        Assigned='No'
        
        #Neighbor of safety valve is searched until a logic node is found
        #neighbor of safety valve
        if Graph.nodes[node]['sub_group_2']=='Safety valves':
            if len(Graph.in_edges(node))==1:
                neighbor=list(nx.all_neighbors(Graph, node))[0] 
                
                #in case of a logic neighbor, the safety valve is assigned to such node
                if Graph.nodes[neighbor]['group'] in ['Vessel', 'Pump', 'Heat exchanger', 'Column', 'Connector', '...']:
                    Assigned='Yes'
                    node_assign=neighbor
                    if type(node_safeguards[node_assign])==list:
                        node_safeguards[node_assign].append('Safety valve')                              
                    else:
                        node_safeguards[node_assign]=['Safety valve'] 
                elif Graph.nodes[neighbor]['Class']=='Pipe tee':
                    Assigned='No'
                elif Graph.nodes[neighbor]['Class']=='Rupture disc':
                    Assigned='No'                    
                    Rupture_disc=neighbor
                else:
                    No_assign.append(node)
                    Assigned='Stop'
            else:
                No_assign.append(node)
                Assigned='Stop'
            
            #if the safety valve could not be assigned to the first neigbor, search hold on 
            while Assigned=='No':
                if len(Graph.in_edges(neighbor))==1:
                    neighbor=list(nx.all_neighbors(Graph, neighbor))[0] 
                    
                    if Graph.nodes[neighbor]['group'] in ['Vessel', 'Pump', 'Heat exchanger', 'Column', 'Connector', '...']:
                        Assigned='Yes'                   
                        node_assign=neighbor
                        
                        #assign the safety vale
                        #if a rupture disc was on the way, it is assigned too
                        if Rupture_disc!='No':
                            Already_assigned.append(Rupture_disc)
                            if type(node_safeguards[node_assign])==list:
                                node_safeguards[node_assign].append('Safety valve')
                                node_safeguards[node_assign].append('Rupture disc')                              
                            else:
                                node_safeguards[node_assign]=['Safety valve'] 
                                node_safeguards[node_assign]=['Rupture disc']
                        else:
                            if type(node_safeguards[node_assign])==list:
                                node_safeguards[node_assign].append('Safety valve')                              
                            else:
                                node_safeguards[node_assign]=['Safety valve']  
                                
                    elif Graph.nodes[neighbor]['Class']=='Pipe tee':
                        Assigned='No'
                    elif Graph.nodes[neighbor]['Class']=='Rupture disc':
                        Assigned='No'                    
                        Rupture_disc=neighbor
                    else:
                        No_assign.append(node)
                        Assigned='Stop'
                else:
                    No_assign.append(node)
                    Assigned='Stop'
         
        #Search function like above
        elif Graph.nodes[node]['Class']=='Rupture disk' and node not in Already_assigned:
            if len(Graph.in_edges(node))==1:
                neighbor=list(nx.all_neighbors(Graph, node))[0]                
                if Graph.nodes[neighbor]['group'] in ['Vessel', 'Pump', 'Heat exchanger', 'Column', 'Connector', '...']:
                    Assigned='Yes'
                    node_assign=neighbor
                    if type(node_safeguards[node_assign])==list:
                        node_safeguards[node_assign].append('Rupture disc')                              
                    else:
                        node_safeguards[node_assign]=['Rupture disc'] 
                elif Graph.nodes[neighbor]['Class']=='Pipe tee':
                    Assigned='No'
                else:
                    No_assign.append(node)
                    Assigned='Stop'
            else:
                No_assign.append(node)
                Assigned='Stop'
                
            while Assigned=='No':
                if len(Graph.in_edges(neighbor))==1:
                    neighbor=list(nx.all_neighbors(Graph, neighbor))[0] 
                    
                    if Graph.nodes[neighbor]['group'] in ['Vessel', 'Pump', 'Heat exchanger', 'Column', 'Connector', '...']:
                        Assigned='Yes'                   
                        node_assign=neighbor
                        if type(node_safeguards[node_assign])==list:
                            node_safeguards[node_assign].append('Rupture disc')                              
                        else:
                            node_safeguards[node_assign]=['Rupture disc']
                                 
                    elif Graph.nodes[neighbor]['Class']=='Pipe tee':
                        Assigned='No'                   
                    else:
                        No_assign.append(node)
                        Assigned='Stop'
                else:
                    No_assign.append(node)
                    Assigned='Stop'
                    
    if No_assign!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='At least one safety valve or rupture disc could not be assigned.'   
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(No_assign)
        Counter_error+=1  


    #ASSIGN FLAP TRAPS TO PUMP
    
    #All Pumps and all flap traps   
    for node_1 in Graph.nodes():
        if Graph.nodes[node_1]['group']=='Pump':
            for node_2 in Graph.nodes():
                print(node_2)
                if Graph.nodes[node_2]['sub_group_2']=='Check valves':
                    
                    #Path between the two selected nodes is searched
                    Paths_from=nx.all_simple_paths(Graph, node_1, node_2)
                    Paths_to=nx.all_simple_paths(Graph, node_2, node_1)
                    
                    #only pipe tees are allowed within the path
                    #distinguish between flap trap in front or follows
                    for Path_from in Paths_from:
                        Check='Yes'
                        for node_3 in Path_from:
                            if Graph.nodes[node_3]['Class']!='Pipe tee' and node_3 not in [node_1, node_2]:
                                Check='No'
                                break
                        #assign to pump
                        if Check=='Yes':
                            if type(node_safeguards[node_1])==list:
                                 node_safeguards[node_1].append('Check valve follows')                               
                            else:
                                node_safeguards[node_1]=['Check valve follows']
                                
                    for Path_to in Paths_to:
                        Check='Yes'
                        for node_3 in Path_to:
                            if Graph.nodes[node_3]['Class']!='Pipe tee' and node_3 not in [node_1, node_2]:
                                Check='No'
                                break
                        if Check=='Yes':
                            if type(node_safeguards[node_1])==list:
                                 node_safeguards[node_1].append('Check valve in front')                               
                            else:
                                node_safeguards[node_1]=['Check valve in front']
                            
                                
    #IDENTIFICATE INERT SYSTEM OF VESSEL
     
    Vessels=[]
    Connectors=[]
    
    #Create new attribute    
    for node in Graph.nodes():
        if Graph.nodes[node]['group'] in ['Vessel', 'Column', '...']:
            Graph.add_node(node, inert='No')                
        else:
            Graph.add_node(node, inert='n.a.')
    
    #find every possible connector, vessel and gas bottle
    for node in Graph.nodes():                        
        if Graph.nodes[node]['group']=='Connector':
            if Graph.nodes[node]['descript'] in ['Inert system', 'Inert System', 'N2', '...']:
                Connectors.append(node)            
        elif Graph.nodes[node]['group'] in ['Vessel', 'Column', '...']:
            Vessels.append(node)
    
    #if a path from Connector to Vessel without another Vessel in between is found, the attribute is given
    for Connector in Connectors:
        for Vessel in Vessels: 
            
            try:
                Path=list(nx.shortest_path(Graph.to_undirected(), Connector, Vessel))
            except nx.NetworkXNoPath:
                Path=[]
                
            if Path!=[]:    
                Check='Yes'
                
                for node in Path:
                    if node in Vessels and node!=Vessel:
                        Check='No'
                        break
                
                if Check=='Yes':
                    Graph.nodes[Vessel]['inert']='Yes'
                    if type(node_safeguards[Vessel])==list:
                         node_safeguards[Vessel].append('Inert system')                               
                    else:
                        node_safeguards[Vessel]=['Inert system']
                            

    #ADD MEASUREMENTS/SIGNALS FOR SAFETY TO SAFEGUARDS
    
    #Measurements
    for node in Graph.nodes():
        if Graph.nodes[node]['measurements'] not in ['n.a.', 'No']:
            for measurement in Graph.nodes[node]['measurements'].split(', '):
                if 'indication' not in measurement and 'registration' not in measurement:
                    if type(node_safeguards[node])==list:
                         node_safeguards[node].append(measurement)                               
                    else:
                        node_safeguards[node]=[measurement]                       
    #Signals
    for node in Graph.nodes():
        if Graph.nodes[node]['signals'] not in ['n.a.', 'No']:
            for Signal in Graph.nodes[node]['signals'].split(', '):
                if 'indication' not in Signal and 'registration' not in Signal:
                    if type(node_safeguards[node])==list:
                         node_safeguards[node].append(Signal)                               
                    else:
                        node_safeguards[node]=[Signal]
    
                
    #FINAL SAFEGUARD ENTRY
    
    for node in node_safeguards.keys():        
        if node_safeguards[node] not in ['n.a.', 'No'] and node in Graph.nodes():#only for nodes with entries
            Final_entry=str(node_safeguards[node])[1:-1].replace("'", "")#Modify the style of entry
            Graph.add_node(node, safeguards=Final_entry)    
    #nx.set_node_attributes(Graph, node_safeguards, Name='safeguards')
            
    
    #REMOVE SAFETY VALVES AND PATHS
    
    Safety_nodes=[]
    
    #select all sefety valves  
    for node_1 in Graph.nodes():
        if Graph.nodes[node_1]['sub_group_2']=='Safety valves':#identificate safety valves
            
            #a valve without an exit only has to be removed, nothing else
            if len(list(Graph.in_edges(node_1)))==1 and len(list(Graph.out_edges(node_1)))==0:            
                Safety_nodes.append(node_1)
                
            #the valve including the exit path to a connector with a certain description is are removed
            elif len(list(Graph.out_edges(node)))>0:
                for node_2 in Graph.nodes():
                    if Graph.nodes[node_2]['group']=='Connector':
                        try:
                            safety_path=nx.shortest_path(Graph, node_1, node_2)
                        except nx.NetworkXNoPath:
                            safety_path=[]
                        if safety_path!=[] and Graph.nodes[node_2]['descript'] in ['Exhaust', 'exhaust', 'Relaxation', 'relaxation', 'reducing', 'Reducing', 'Pressure', 'pressure']:
                            Safety_nodes=Safety_nodes+safety_path
                                
    Graph.remove_nodes_from(Safety_nodes)   


    #REMOVE UNNECESSARY PIPE TEES
    
    another_cycle='Yes' 
    
    while another_cycle=='Yes':
        list_pipe_tees=[]
        for node in Graph.nodes():
            all_neighbors=list((nx.all_neighbors(Graph, node)))
            neighbors=list((nx.neighbors(Graph, node)))
            if Graph.nodes[node]['Class']=='Pipe tee':
                if len(all_neighbors)==2 and len(neighbors)==1:
                    list_pipe_tees.append(node)
                    FromID=all_neighbors[0]
                    ToID=all_neighbors[1]
                    C=nx.get_edge_attributes(Graph, 'Class')[(FromID, node)]
                    C_sub=nx.get_edge_attributes(Graph, 'Sub_class')[(FromID, node)]
                    Group=nx.get_edge_attributes(Graph, 'group')[(FromID, node)]
                    Sub_group=nx.get_edge_attributes(Graph, 'sub_group')[(FromID, node)]
                    M=nx.get_edge_attributes(Graph, 'material')[(FromID, node)]
                    D=nx.get_edge_attributes(Graph, 'diameter')[(FromID, node)]
                    C_pipe=nx.get_edge_attributes(Graph, 'pipe_class')[(FromID, node)]
                    I=nx.get_edge_attributes(Graph, 'insulation')[(FromID, node)]    
                    HC=nx.get_edge_attributes(Graph, 'heated_cooled')[(FromID, node)]
                    PS=nx.get_edge_attributes(Graph, 'Part_simu')[(FromID, node)]
                    Subs=nx.get_edge_attributes(Graph, 'subs')[(FromID, node)]                    
                    Simu_P=nx.get_edge_attributes(Graph, 'P_operation')[(FromID, node)]
                    Simu_unit_P=nx.get_edge_attributes(Graph, 'P_operation_unit')[(FromID, node)]                    
                    Simu_T=nx.get_edge_attributes(Graph, 'T_operation')[(FromID, node)]
                    Simu_unit_T=nx.get_edge_attributes(Graph, 'T_operation_unit')[(FromID, node)]  
                    Simu_MF=nx.get_edge_attributes(Graph, 'Mass_flow')[(FromID, node)]
                    Simu_unit_MF=nx.get_edge_attributes(Graph, 'Mass_flow_unit')[(FromID, node)] 
                    Simu_VF=nx.get_edge_attributes(Graph, 'Volume_flow')[(FromID, node)]
                    Simu_unit_VF=nx.get_edge_attributes(Graph, 'Volume_flow_unit')[(FromID, node)]                    
                    Simu_density=nx.get_edge_attributes(Graph, 'Density')[(FromID, node)]
                    Simu_unit_density=nx.get_edge_attributes(Graph, 'Density_unit')[(FromID, node)]  
                    Sol_frac=nx.get_edge_attributes(Graph, 'sol_frac')[(FromID, node)]
                    Liq_frac=nx.get_edge_attributes(Graph, 'liq_frac')[(FromID, node)] 
                    Vap_frac=nx.get_edge_attributes(Graph, 'liq_frac')[(FromID, node)]                     
                     
                    Graph.add_edge(FromID, ToID, Class=C, Sub_class=C_sub, material=M,
                           diameter=D, pipe_class=C_pipe, insulation=I, heated_cooled=HC, group=Group, sub_group=Sub_group,
                           Part_simu=PS, subs=Subs,
                           P_operation=Simu_P, P_operation_unit=Simu_unit_P, 
                           T_operation=Simu_T, T_operation_unit=Simu_unit_T,
                           Mass_flow=Simu_MF, Mass_flow_unit=Simu_unit_MF, 
                           Volume_flow=Simu_VF, Volume_flow_unit=Simu_unit_VF,
                           Density=Simu_density, Density_unit=Simu_unit_density,
                           sol_frac=Sol_frac, liq_frac=Liq_frac, vap_frac=Vap_frac)               
                    
        if len(list_pipe_tees)!=0:
            Graph.remove_nodes_from(list_pipe_tees)   
        else:
            another_cycle='No'     

    
    #SAVE
    
    Results_workbook.save(Path_results)  
    nx.write_graphml(Graph, Path_Graph, encoding='utf-8', prettyprint=True, infer_numeric_types=False)        
    return [Graph]  

                         