# -*- coding: utf-8 -*-
"""
Created on Wed Dec  8 11:22:45 2021

@Author: Tim Holtermann, TU Dortmund, BCI AG Apparatedesign
"""

def Dexpi2graph(DEXPI_path, Path_graph, Path_results, Path_results_template):
    
    import networkx as nx
    import xml.etree.ElementTree as ET
    import nltk
    import sys
    from openpyxl import load_workbook
    
    mytree = ET.parse(DEXPI_path)#load DEXPI-File    
    myroot = mytree.getroot()         
    graph=nx.DiGraph()# create directed graph   

    
    #WORKBOOKS    
   
    Template_workbook=load_workbook(Path_results_template)
    Error_log=Template_workbook['Error_log']
    Counter_error=2
    max_col=Error_log.max_column
    Header_index={}  
    for i in range(1, max_col + 1): 
        Header = Error_log.cell(row = 1, column = i).value
        Header_index[Header]=i

    ID_list=Template_workbook['ID_list']
    Counter_ID_list=2
    max_col=ID_list.max_column
    Header_index_ID_list={}  
    for i in range(1, max_col + 1): 
        Header = ID_list.cell(row = 1, column = i).value
        Header_index_ID_list[Header]=i

    
    #EQUIPMENT
    
    i=0
    
    for equipment in myroot.findall('Equipment'):#all equipments
    
        #predefine the optional attributes
        P1='n.a.'
        unit_P1='n.a.'
        unit_P2='n.a.'
        P2='n.a.'   
        T1='n.a.'
        unit_T1='n.a.'  
        unit_T2='n.a.'  
        T2='n.a.' 
        V='n.a.'
        unit_V='n.a.'
        I='n.a.'
        CH='n.a.'
        L='n.a.'
        M='n.a.'
        R='n.a.'
        N_owner='n.a.'
        Descript='n.a.'
        Number='n.a.'
        
        #read
        ID=equipment.get('ID')#for nozzle assignment
        ID_saved=ID
        N=equipment.get('TagName')
        if N=="":
            N='n.d.'
            
        #read equipment coordinates.
        Position = equipment.findall('Position/Location')
        x=(Position[0].get('X').replace(',', '.'))
        y=(Position[0].get('Y').replace(',', '.'))
        
        #Read all equipment attributes
        for equipment_attributes in equipment.findall('GenericAttributes/GenericAttribute'):
            if equipment_attributes.get('Name')=='CLASS':#selecting the class
                C=equipment_attributes.get('Value')
                
            elif equipment_attributes.get('Name')=='SUB_CLASS':#selecting the sub class
                C_sub=equipment_attributes.get('Value')
                
            elif equipment_attributes.get('Name')=='VPE_PRESSURE_DESIGN_MAX':#selecting max design pressure
                if len(nltk.word_tokenize(equipment_attributes.get('Value')))==2:
                    P1=(nltk.word_tokenize(equipment_attributes.get('Value'))[0].replace(',', '.'))
                    unit_P1=nltk.word_tokenize(equipment_attributes.get('Value'))[1]
                elif equipment_attributes.get('Value')=='': 
                    P1=''
                    unit_P1=''
                else:
                    sys.exit('Invalid entry in Equipment properties1') 
                if P1=="" and unit_P1=="":
                    P1='n.d.'
                    unit_P1='n.d.'
                    
            elif equipment_attributes.get('Name')=='VPE_PRESSURE_DESIGN_MIN':#selecting min design pressure
                if len(nltk.word_tokenize(equipment_attributes.get('Value')))==2:
                    P2=nltk.word_tokenize(equipment_attributes.get('Value'))[0].replace(',', '.')#first token of dexpi string is selected
                    unit_P2=nltk.word_tokenize(equipment_attributes.get('Value'))[1]
                elif equipment_attributes.get('Value')=='': 
                    P2=''
                    unit_P2=''
                else:
                    sys.exit('Invalid entry in Equipment properties2') 
                if P2=="" and unit_P2=="":
                    P2='n.d.'
                    unit_P2='n.d.'
                    
            elif equipment_attributes.get('Name')=='VPE_TEMP_DESIGN_MAX':#selecting max design temperature
                if len(nltk.word_tokenize(equipment_attributes.get('Value')))==2:            
                    T1=nltk.word_tokenize(equipment_attributes.get('Value'))[0].replace(',', '.')
                    unit_T1=nltk.word_tokenize(equipment_attributes.get('Value'))[1]
                elif equipment_attributes.get('Value')=='': 
                    T1=''
                    unit_T1=''
                else:
                    sys.exit('Invalid entry in Equipment properties3') 
                if T1=="" and unit_T1=="":
                    T1='n.d.'
                    unit_T1='n.d.'
                    
            elif equipment_attributes.get('Name')=='VPE_TEMP_DESIGN_MIN':#selecting min design temperature
                if len(nltk.word_tokenize(equipment_attributes.get('Value')))==2:  
                    T2=nltk.word_tokenize(equipment_attributes.get('Value'))[0].replace(',', '.')
                    unit_T2=nltk.word_tokenize(equipment_attributes.get('Value'))[1]
                elif equipment_attributes.get('Value')=='': 
                    T2=''
                    unit_T2=''
                else:
                    sys.exit('Invalid entry in Equipment properties4') 
                if T2=="" and unit_T2=="":
                    T2='n.d.'
                    unit_T2='n.d.'
                
            elif equipment_attributes.get('Name')=='VPE_TNK_VOL_BRUTTO':#selecting volume
                if len(nltk.word_tokenize(equipment_attributes.get('Value')))==2: 
                    V=nltk.word_tokenize(equipment_attributes.get('Value'))[0].replace(',', '.')
                    unit_V=nltk.word_tokenize(equipment_attributes.get('Value'))[1]
                elif equipment_attributes.get('Value')=='': 
                    V=''
                    unit_V=''
                else:
                    sys.exit('Invalid entry in Equipment properties5') 
                if V=="" and unit_V=="":
                    V='n.d.'
                    unit_V='n.d.'
                    
            elif equipment_attributes.get('Name')=='VPE_MAT_PARTS_MEDIA_CONTACT':#selecting material in contact with medium
                M=equipment_attributes.get('Value')
                if M=="":
                    M='n.d.'                    
                    
            elif equipment_attributes.get('Name')=='INSULATION':#selecting insulation (Yes/No)
                I=equipment_attributes.get('Value') 
                if I=="":
                    I='n.d.'
                
            elif equipment_attributes.get('Name')=='COOLING_HEATING_SYSTEM':#selecting cooling/heating system
                CH=equipment_attributes.get('Value')
                if CH=="":
                    CH='No'
                
        #Add equipment as node with attributes to graph
        graph.add_node(ID, DEXPI_ID=ID, PID_label=N, Class=C, Sub_class=C_sub, X=x, Y=y, 
                   P_max_design=P1, P_max_design_unit=unit_P1, P_min_design=P2, P_min_design_unit=unit_P2, 
                   T_max_design=T1, T_max_design_unit=unit_T1, T_min_design=T2, T_min_design_unit=unit_T2,
                   V=V, V_unit=unit_V,
                   material=M, insulation=I, C_H_system=CH, location=L, request=R, 
                   nozzle_owner=N_owner,
                   number=Number, descript=Descript)#create node with attributes             
        
        #Read all nozzles
        for equipment_nozzle in equipment.findall('Nozzle'):
            
            #predefine the optional attributes
            P1='n.a.'
            unit_P1='n.a.'
            unit_P2='n.a.'
            P2='n.a.'   
            T1='n.a.'
            unit_T1='n.a.'  
            unit_T2='n.a.'  
            T2='n.a.' 
            V='n.a.'
            unit_V='n.a.'
            I='n.a.'
            CH='n.a.'
            L='n.a.'
            M='n.a.'
            R='n.a.'
            N_owner='n.a.'
            Descript='n.a.'
            Number='n.a.'
            
            N_owner=ID_saved#take the ID from equipment before
            ID=equipment_nozzle.get('ID')
            N=equipment_nozzle.get('TagName')
            if N=="":
                N='n.d.'
            
            #read equipment coordinates.
            Position = equipment_nozzle.findall('Position/Location')
            x=Position[0].get('X').replace(',', '.') 
            y=Position[0].get('Y').replace(',', '.')
            
            #Read all equipment attributes
            for equipment_nozzle_attribute in equipment_nozzle.findall('GenericAttributes/GenericAttribute'):
                if equipment_nozzle_attribute.get('Name')=='CLASS':#selecting the class
                    C=equipment_nozzle_attribute.get('Value')            
                elif equipment_nozzle_attribute.get('Name')=='SUB_CLASS':#selecting the class
                    C_sub=equipment_nozzle_attribute.get('Value') 
                    
            graph.add_node(ID, DEXPI_ID=ID, PID_label=N, Class=C, Sub_class=C_sub, X=x, Y=y, 
                   P_max_design=P1, P_max_design_unit=unit_P1, P_min_design=P2, P_min_design_unit=unit_P2, 
                   T_max_design=T1, T_max_design_unit=unit_T1, T_min_design=T2, T_min_design_unit=unit_T2,
                   V=V, V_unit=unit_V,
                   material=M, insulation=I, C_H_system=CH, location=L, request=R, 
                   nozzle_owner=N_owner,
                   number=Number, descript=Descript)
                       
    #MSR
    
    for PIF in myroot.findall('ProcessInstrumentationFunction'):#all MSR
    
        #predefine the optional attributes
        P1='n.a.'
        unit_P1='n.a.'
        unit_P2='n.a.'
        P2='n.a.'   
        T1='n.a.'
        unit_T1='n.a.'  
        unit_T2='n.a.'  
        T2='n.a.' 
        V='n.a.'
        unit_V='n.a.'
        I='n.a.'
        CH='n.a.'
        L='n.a.'
        M='n.a.'
        R='n.a.'
        N_owner='n.a.'
        Descript='n.a.'
        Number='n.a.'
        
        #Read
        ID=PIF.get('ID')
        N=PIF.get('TagName')
        if N=="":
            N='n.d.'
        
        #Read equipment coordinates
        Position = PIF.findall('Position/Location')
        x=Position[0].get('X').replace(',', '.')
        y=(Position[0].get('Y').replace(',', '.'))
        
        #Read all attributes
        for PIF_attributes in PIF.findall('GenericAttributes/GenericAttribute'):
            
            if PIF_attributes.get('Name')=='CLASS':#selecting the class
                C=PIF_attributes.get('Value')
                
            if PIF_attributes.get('Name')=='PCE_CAT_FUNC':#selecting the type of request
                R=PIF_attributes.get('Value')
                
            elif PIF_attributes.get('Name')=='SUB_CLASS':#selecting the sub class
                C_sub=PIF_attributes.get('Value')
                
            elif PIF_attributes.get('Name')=='LOCATION':#selecting location
                L=PIF_attributes.get('Value')
        
        #Add MSR-unit as node with attributes to graph            
        graph.add_node(ID, DEXPI_ID=ID, PID_label=N, Class=C, Sub_class=C_sub, X=x, Y=y, 
                       P_max_design=P1, P_max_design_unit=unit_P1, P_min_design=P2, P_min_design_unit=unit_P2, 
                       T_max_design=T1, T_max_design_unit=unit_T1, T_min_design=T2, T_min_design_unit=unit_T2,
                       V=V, V_unit=unit_V,
                       material=M, insulation=I, C_H_system=CH, location=L, request=R, 
                       nozzle_owner=N_owner,
                       number=Number, descript=Descript)
    
    #PIPING COMPONENTS
    
    j=1         
    k=101
       
    for piping_component in myroot.findall('PipingNetworkSystem//PipingComponent'):#select all piping components
    
        #predefine optional attributes
        P1='n.a.'
        unit_P1='n.a.'
        unit_P2='n.a.'
        P2='n.a.'   
        T1='n.a.'
        unit_T1='n.a.'  
        unit_T2='n.a.'  
        T2='n.a.' 
        V='n.a.'
        unit_V='n.a.'
        I='n.a.'
        CH='n.a.'
        L='n.a.'
        M='n.a.'
        R='n.a.'
        N_owner='n.a.'
        Descript='n.a.'
        Number='n.a.'
        
        #Read
        ID=piping_component.get('ID')
        N=piping_component.get('TagName')
        if N=="":
            N='n.d.'
            
        #Read equipment coordinates
        Position = piping_component.findall('Position/Location')
        x=(Position[0].get('X').replace(',', '.'))
        y=(Position[0].get('Y').replace(',', '.'))
        
        #in case of a pipe tee
        if piping_component.get('ComponentClass')=='Pipe tee': 
               
            C=piping_component.get('ComponentClass')
            N='T'+str(j)#P&ID name does not exist, so name by consecutive numbers
            j+=1
        
        #in case of an in/-outlet    
        elif piping_component.get('ComponentClass')=='Arrow' or piping_component.get('ComponentClass')=='Flow in pipe connector symbol' or piping_component.get('ComponentClass')=='Flow out pipe connector symbol':
            
            #Read attributes
            for Attribute in piping_component.findall('GenericAttributes/GenericAttribute'):
                if Attribute.get('Name')=='CLASS':
                    C=Attribute.get('Value')
                elif Attribute.get('Name')=='SUB_CLASS':
                    C_sub=Attribute.get('Value')
                elif Attribute.get('Name')=='DESCRIPT':
                    Descript=Attribute.get('Value') 
                    if Descript=="":
                        Descript='n.d.'
                elif Attribute.get('Name')=='INLET_OUTLET_NO':
                    Number=Attribute.get('Value') 
                    if Number=="":
                        Number='n.d.'
            
            #Special labeling of in-outlet
            if Number not in ['n.d.', 'n.a.']:
                N='C'+Number
            else:
                print('At least one In-outlet is not labeled with a number.')
                N='C'+str(k)
                k+=1
                        
        #other piping components    
        else:            
            #Read attributes
            for piping_component_attribute in piping_component.findall('GenericAttributes/GenericAttribute'):
                if piping_component_attribute.get('Name')=='CLASS':
                    C=piping_component_attribute.get('Value')
                elif piping_component_attribute.get('Name')=='SUB_CLASS':
                    C_sub=piping_component_attribute.get('Value')
                                
        #Add piping component node with attributes to graph
        graph.add_node(ID, DEXPI_ID=ID, PID_label=N, Class=C, Sub_class=C_sub, X=x, Y=y, 
                   P_max_design=P1, P_max_design_unit=unit_P1, P_min_design=P2, P_min_design_unit=unit_P2, 
                   T_max_design=T1, T_max_design_unit=unit_T1, T_min_design=T2, T_min_design_unit=unit_T2,
                   V=V, V_unit=unit_V,
                   material=M, insulation=I, C_H_system=CH, location=L, request=R, 
                   nozzle_owner=N_owner,
                   number=Number, descript=Descript)        

        
    #PIPING CONNECTIONS
    
    i=1
    nodes_from_nothing=[]
    nodes_to_nothing=[]
    nodes_not_registrated=[]
    
    for PNS in myroot.findall('PipingNetworkSystem/PipingNetworkSegment'):
        
        #identificate connection
        for connection in PNS.findall('Connection'):
            FromID=connection.get('FromID')
            ToID=connection.get('ToID')
            
            #Connection only added if there is a start and end point
            if FromID=="" or FromID==None:
               nodes_from_nothing.append(ToID) 
            elif ToID=="" or ToID==None:
               nodes_to_nothing.append(FromID)     
            elif FromID not in list(graph.nodes()):
               nodes_not_registrated.append(FromID) 
            elif ToID not in list(graph.nodes()):
               nodes_not_registrated.append(ToID) 
            else:
                #predefine optional parameter
                M='n.a.'
                D='n.a.'
                C_pipe='n.a.'

                #Read attributes
                for piping_attribute in PNS.findall('GenericAttributes/GenericAttribute'):
                    
                    if piping_attribute.get('Name')=='CLASS':#selecting the class
                        C=piping_attribute.get('Value')
                        
                    elif piping_attribute.get('Name')=='SUB_CLASS':#selecting the sub class
                        C_sub=piping_attribute.get('Value')
                        
                    elif piping_attribute.get('Name')=='VPE_MAT_MAIN_MATERIAL':#selecting the material
                        M=piping_attribute.get('Value')
                        if M=="":
                            M='n.d.'
                        
                    elif piping_attribute.get('Name')=='NOMINAL_DIAMETER':#selecting Diameter
                        D=piping_attribute.get('Value')
                        if D=="":
                            D='n.d.'
                        
                    elif piping_attribute.get('Name')=='MAT_INAME':#selecting pipe class
                        C_pipe=piping_attribute.get('Value')
                        if C_pipe=="":
                            C_pipe='n.d.'
                                                                              
                #Add piping connection as edge with attributes to graph
                graph.add_edge(FromID, ToID, Class=C, Sub_class=C_sub, material=M,
                           diameter=D, pipe_class=C_pipe)#Adding connection as edge to graph   
 
                              
    #MSR CONNECTIONS
    
    for InfoFlow in myroot.findall('ProcessInstrumentationFunction/InformationFlow'):
        
        for connection in InfoFlow.findall('Connection'):
            FromID=connection.get('FromID')
            ToID=connection.get('ToID')
            
            #Connection only added if there is a start and end point
            if FromID=="" or FromID==None:
               nodes_from_nothing.append(ToID) 
            elif ToID=="" or ToID==None:
               nodes_to_nothing.append(FromID)     
            elif FromID not in list(graph.nodes()):
               nodes_not_registrated.append(FromID) 
            elif ToID not in list(graph.nodes()):
               nodes_not_registrated.append(ToID) 
            else:            
                #predefine optional parameter
                M='n.a.'
                D='n.a.'
                C_pipe='n.a.'
                
                #Read attributes
                for attributes in InfoFlow.findall('GenericAttributes/GenericAttribute'):
                    
                    if attributes.get('Name')=='CLASS':#selecting the class
                        C=attributes.get('Value')
                        
                    elif attributes.get('Name')=='SUB_CLASS':#selecting the sub class
                        C_sub=attributes.get('Value')
                
                #Add MSR connection as edge to graph
                graph.add_edge(FromID, ToID, Class=C, Sub_class=C_sub, material=M,
                           diameter=D, pipe_class=C_pipe)#Adding connection as edge to graph
    

    #Add warning(s) to the error log    
    
    if nodes_from_nothing!=[]:       
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='There is at least one node without a source. Please make sure it is correct.'
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_from_nothing)[1:-1]
        Counter_error+=1
    if nodes_to_nothing!=[]:       
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='There is at least one node without a destination. Please make sure it is correct.' 
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_to_nothing)[1:-1]
        Counter_error+=1 
    if nodes_not_registrated!=[]:      
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='At least one exported edge contains a node that was not registrated before.'
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_not_registrated)[1:-1]
        Counter_error+=1
    
    
    ###################################################################################################################      
    ### PROCESS DATA ################################################################################################
    ###################################################################################################################
    
        
    #REMOVE EMPTY NODES
    
    Empty_1='No'
    Empty_2='No'
    
    for node in graph.nodes():
        if node == "":
            Empty_1='Yes'
        if node == None:
            Empty_2='Yes'
            
    if Empty_1=='Yes':
        graph.remove_node("")    
    if Empty_2=='Yes':
        graph.remove_node(None)
    
    
    #PROCESS NOZZLES
    
    Nozzles=[]
    
    for edge in graph.edges():
        if graph.edges[edge]['Class'] in ['Piping', 'Heat transfer medium', 'Process connection line']:
            graph.add_edge(edge[0], edge[1], nozzle='No')
        else:
            graph.add_edge(edge[0], edge[1], nozzle='n.a.')    
    
    #create an edge from or to the nozzle owner
    for node in graph.nodes():
        if graph.nodes[node]['Class']=='Nozzle':
            if len(list(graph.in_edges(node)))==1 and len(list(graph.out_edges(node)))==0:
                FromID=list(nx.all_neighbors(graph, node))[0]
                C=nx.get_edge_attributes(graph, 'Class')[(FromID, node)]
                C_sub=nx.get_edge_attributes(graph, 'Sub_class')[(FromID, node)]
                M=nx.get_edge_attributes(graph, 'material')[(FromID, node)]
                D=nx.get_edge_attributes(graph, 'diameter')[(FromID, node)]
                C_pipe=nx.get_edge_attributes(graph, 'pipe_class')[(FromID, node)]
                graph.add_edge(FromID, graph.nodes[node]['nozzle_owner'], Class=C, Sub_class=C_sub, material=M,
                           diameter=D, pipe_class=C_pipe, nozzle=node)
                
            if len(list(graph.out_edges(node)))==1 and len(list(graph.in_edges(node)))==0:
                ToID=list(nx.neighbors(graph, node))[0]
                C=nx.get_edge_attributes(graph, 'Class')[(node, ToID)]
                C_sub=nx.get_edge_attributes(graph, 'Sub_class')[(node, ToID)]
                M=nx.get_edge_attributes(graph, 'material')[(node, ToID)]
                D=nx.get_edge_attributes(graph, 'diameter')[(node, ToID)]
                C_pipe=nx.get_edge_attributes(graph, 'pipe_class')[(node, ToID)]               
                graph.add_edge(graph.nodes[node]['nozzle_owner'], ToID, Class=C, Sub_class=C_sub, material=M,
                           diameter=D, pipe_class=C_pipe, nozzle=node)                

            Nozzles.append(node)
                                
    graph.remove_nodes_from(Nozzles)
    
        
    #CONVERT NODES LIKE PIPING EQUIPMENT AND HOSE IN TO EDGES
    
    remove=[]
    nodes_problem=[]
    pipe_attributes={'Piping with conduit':{'Insulation':'No', 'Heated/cooled':'No'},
                     'Piping insulated':{'Insulation':'Yes', 'Heated/cooled':'No'},
                     'Piping heated or cooled':{'Insulation':'No', 'Heated/cooled':'Yes'},
                     'Piping, heating or cooled insulated':{'Insulation':'Yes', 'Heated/cooled':'Yes'}}#preparation for creating new edge with new attributes (dict to avoid repeating script)
    
    #Add following new attributes to the already existing nodes
    for edge in graph.edges():
        if graph.edges[edge]['Class']=='Piping':
            graph.add_edge(edge[0], edge[1], insulation='No', heated_cooled='No')
        else:
            graph.add_edge(edge[0], edge[1], insulation='n.a.', heated_cooled='n.a.')
    
    # Select all relevant nodes
    for node in graph.nodes():
        if graph.nodes[node]['Class'] in ['Hose', 'Pipe equipment']:
            
            #checking the right connection format, save node and its neighbors
            if len(list((graph.in_edges(node))))==1 and len(list((graph.out_edges(node))))==1:
                remove.append(node)
                all_neighbors=list((nx.all_neighbors(graph, node)))
                FromID=all_neighbors[0]
                ToID=all_neighbors[1]
                
                #In case of a pipe equipment the attributes of the edge in front can be taken over
                if graph.nodes[node]['Class']=='Pipe equipment': 
                    C=nx.get_edge_attributes(graph, 'Class')[(FromID, node)]
                    C_sub=nx.get_edge_attributes(graph, 'Sub_class')[(FromID, node)]
                    M=nx.get_edge_attributes(graph, 'material')[(FromID, node)]
                    D=nx.get_edge_attributes(graph, 'diameter')[(FromID, node)]
                    C_pipe=nx.get_edge_attributes(graph, 'pipe_class')[(FromID, node)]
    
                    #the remaining additional attributes depend on the sub class of pipe equipment (saved in the dict above)
                    for sub_class in pipe_attributes:  
                        if graph.nodes[node]['Sub_class']==sub_class:                
                            graph.add_edge(FromID, ToID, Class=C, Sub_class=C_sub, material=M,
                                       diameter=D, pipe_class=C_pipe,
                                       insulation=pipe_attributes[sub_class]['Insulation'],
                                       heated_cooled=pipe_attributes[sub_class]['Heated/cooled'])
                            
                #in case of a hose, attributes can not be taken from pipe in front cause of having own attriubtes
                elif graph.nodes[node]['Class']=='Hose':
                    #predefine optional parameter
                    M='n.a.'
                    D='n.a.'
                    C_pipe='n.a.'                 
                    
                    #Read hose attributes
                    C=graph.nodes[node]['Class']
                    C_sub=graph.nodes[node]['Sub_class']
    
                    graph.add_edge(FromID, ToID, Class=C, Sub_class=C_sub, material=M,
                               diameter=D, pipe_class=C_pipe,
                               insulation='No',
                               heated_cooled='No')
    
            #if the node is not connected in the expected way, a warning is shown
            else:
                nodes_problem.append(node)
     
    if nodes_problem!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='Problems in converting hose and piping components. At least one node is connected in an unexpected way.'
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_problem[1:-1])
        Counter_error+=1
        
    graph.remove_nodes_from(remove)#remove old nodes for pipe equipment and nodes from graph
    

    #EDGE GROUPS

    dict_group={'Medium transfer':['Piping', 'Heat transfer medium', 'Hose'],
                'Signal':['Signal line'],
                'Process connection':['Process connection line']}#group

    #create following attributes
    for edge in graph.edges():
        graph.add_edge(edge[0], edge[1], group='n.a.')
        graph.add_edge(edge[0], edge[1], sub_group='n.a.')
            
    #Class
    for name_group in dict_group:
        for edge in graph.edges():
            if graph.edges[edge]['Class'] in dict_group[name_group]:
                graph.add_edge(edge[0], edge[1], group=name_group)#overwrite attribute
     
    #Sub class
    for edge in graph.edges():
        if graph.edges[edge]['Class']=='Heat transfer medium':
            graph.add_edge(edge[0], edge[1], sub_group='Heat transfer pipe')#overwrite attribute 
        if graph.edges[edge]['Sub_class']=='Main pipe':
            graph.add_edge(edge[0], edge[1], sub_group='Main pipe')#overwrite attribute         
        if graph.edges[edge]['Sub_class']=='Secondary pipe':
            graph.add_edge(edge[0], edge[1], sub_group='Secondary pipe')#overwrite attribute
        if graph.edges[edge]['Class']=='Hose':
            graph.add_edge(edge[0], edge[1], sub_group='Hose')#overwrite attribute                   
    
    
    #NODE GROUPS
    
    dict_group={'Vessel':['Vessel', 'Vessel with two Diameters', 'Spherical vessel', 'Vessel with dome', 'Vessel, general', 'Silo', 'Gas cylinder', 'Basin', 'Barrel', 'Tank', 'Vessel with agigator', 'Bag', 'Container'],
                'Column':['Column'],
                'Shaping machines':['Vertical shaping machine', 'Horizontal shaping machine'],
                'Crushing/Grinding':['Crushing maschine', 'Mill'],
                'Dryer':['Dryer'],
                'Centrifuge':['Centrifuge'],
                'Separator':['Separator'],
                'Sieving':['Basket band and screening machine', 'Sieving machine'],
                'Mixer/Kneader':['Kneader', 'Mixing pipe', 'Rotating mixer', 'Static mixer'],
                'Fittings':['Flange', 'Steam trap', 'Orifice plate', 'Flame arrestor', 'Fire protection flap', 'Rupture disk'],
                'Valves':['Valve (general)', 'Valve, angle type (general)', 'Valve, three way type (general)', 'Valve, globe type', 'Valve, angle globe type', 'Valve, three way globe type', 'Valve, ball type', 'Valve, angle ball type', 'Valve, three way ball type', 'Valve, gate type', 'Valve, butterfly type (Form 1)', 'Valve, butterfly type (Form 2)', 'Valve, needle type', 'Diaphragm valve', 'Plug clock', 'Plug valve',
                          'Fire protection flap', 'Breather flap', 'Breather valve', 'Safety valve, angled type', 'Safety valve',
                          'Check valve', 'Flap trap (form 1)', 'Flap trap (form 2)', 'Check valve globe type', 'Check valve angled globe type', 'Check valve angled',
                          'Angle check valve', 'Angle globe valve', 'Angle valve, general', 'Ball valve', 'butterfly valve', 'Butterfly valve', 'Float valve', 'Gate valve', 'Globe valve', 'Pressure reducing valve', 'Safety angle valve', 'Swing check valve', 'Three-way ball valve', 'Three-way globe valve', 'Three-way valve, general', 'Valve, general', 'Valve, three-way ball type', 'Valve, three-way globe type', 'Valve, three-way type (general)', 'Airtight butterfly valve'],
                'Filter':['Filter', 'Band Filter', ' Ion exchange filter', 'Air filter', 'Biological filter', 'Filter press', 'Fluid filter', 'Gas filter', 'Liquid rotary filter'],
                'Pump':['Fluid pump', 'Liquid pump', 'Liquid jet pump'],
                'Compressor':['Compressor', 'Ejector compressor', 'Vacuum pump', 'Vakuum pump', 'Jet vacuum pump', 'Jet vakuum pump'],
                'Heat exchanger':['Heat exchanger', 'Heat exchanger ', 'Heat exchanger, detailed', 'Spiral type heat exchanger','Heat exchanger', 'detailed, Tube bundle with U-tubes', 'Electric Heaters', 'Facility for heating or cooling'],
                'Pipe tee':['Pipe tee'],
                'Connector':['In-outlet', 'Arrow', 'Flow in pipe connector symbol', 'Flow out pipe connector symbol'],
                'MSR':['PCE Request']}#groups
    
    #create following attributes
    for node in graph.nodes():
        graph.add_node(node, group='n.a.')

    #sorting into groups by the dict (node class has to match with one of the listed classes under a specific group name)
    for name_group in dict_group:
        for node in graph.nodes():
            if graph.nodes[node]['Class'] in dict_group[name_group]:
                graph.add_node(node, group=name_group)#overwrite attribute 


    #SUB GROUPS
    
    dict_sub_group={'Valves':{'Valves (safety)':{'Safety valve':'all', 'Safety valve, angled type':'all', 'Safety angle valve':['Spring loaded', 'General'], 'Check valve':'all', 'Flap trap (form 1)':'all', 'Flap trap (form 2)':'all', 'Check valve globe type':'all', 'Check valve angled globe type':'all', 'Check valve angled':'all', 'Breather valve':'all', 'Breather flap':'all'},
                              'Valves (operation)':'All other'},
                    'Heat exchanger':{'Heat exchanger (heating)':{'Evaporator':'General', 'Film Evaporator':'General', 'Electric Heaters':'General'},
                                      'Heat exchanger (cooling)':'Not yet possible'}}#sub groups 
    
    dict_sub_group_2={'Heat exchanger':{'Heat exchanger (heating)':{'Heater':'Not yet possible',
                                                                      'Evaporator':{'Evaporator':'General', 'Film Evaporator':'General'}},
                                          'Heat exchanger (cooling)':{'Cooler':'Not yet possible',
                                                                      'Condenser':'Not yet possible'}},
                      'Valves':{'Valves (safety)':{'Safety valves':{'Safety valve':'all', 'Safety valve, angled type':'all', 'Safety angle valve':['Spring loaded', 'General']},
                                                   'Check valves':{'Check valve':'all', 'Flap trap (form 1)':'all', 'Flap trap (form 2)':'all', 'Check valve globe type':'all', 'Check valve angled globe type':'all', 'Check valve angled':'all'},
                                                   'Breather valves':{'Breather valve':'all', 'Breather flap':'all'}}}}


    #create following attributes
    for node in graph.nodes():
        graph.add_node(node, sub_group='n.a.')
        graph.add_node(node, sub_group_2='n.a.')           

    no_sub_group=[]#List to collect the nodes which are not sorted after sorting
    
    #sorting into sub groups by the dict
    for node in graph.nodes():
        grouped='No'
        for group in dict_sub_group:#regarding all groups  
            if graph.nodes[node]['group']==group:#select nodes belongs to the regarded group
                for sub_group in dict_sub_group[group]:#regarding all sub groups
                
                    if dict_sub_group[group][sub_group]=='All other':#everything else
                        graph.add_node(node, sub_group=sub_group)
                        grouped='Yes'
                        break
                        
                    else:
                        for Class in dict_sub_group[group][sub_group]:
                            if Class==graph.nodes[node]['Class']:                            
                                sub_classes=dict_sub_group[group][sub_group][Class]
                                
                                #Case of only one sub class (entry as string) or the signal word "all"
                                if type(sub_classes)==str:                                                       
                                    if sub_classes==graph.nodes[node]['Sub_class'] or sub_classes=='all':    
                                        graph.add_node(node, sub_group=sub_group)
                                        grouped='Yes'
                                        break
                                        
                                #Case of more than one sub classes (entry as list)        
                                elif type(sub_classes)==list:
                                    for Subclass in sub_classes:
                                        if Subclass==graph.nodes[node]['Sub_class']:
                                            graph.add_node(node, sub_group=sub_group)
                                            grouped='Yes'
                                            break
                                        
            #To break out of all loops except the last to start the next sorting                            
                            if grouped=='Yes':
                                break
                    if grouped=='Yes':
                        break
            if grouped=='Yes':
                break
        
        #If no sub group was found for the node (Requirement is that a potential sub group must exist)    
        if grouped=='No' and graph.nodes[node]['group'] in dict_sub_group.keys():
            no_sub_group.append(node)                                        
    
    
    #sorting into further sub groups by the dict
    for node in graph.nodes():
        grouped='No'
        for group in dict_sub_group_2:#regarding all groups  
            if graph.nodes[node]['group']==group:#select nodes belongs to the regarded group
                for sub_group in dict_sub_group_2[group]:#regarding all sub groups
                    for sub_group_2 in dict_sub_group_2[group][sub_group]:#regarding all sub2 groups                
                        if dict_sub_group_2[group][sub_group][sub_group_2]=='All other':#everything else
                            graph.add_node(node, sub_group_2=sub_group_2)
                            grouped='Yes'
                            break
                        
                        else:
                            for Class in dict_sub_group_2[group][sub_group][sub_group_2]:
                                if Class==graph.nodes[node]['Class']:                            
                                    sub_classes=dict_sub_group_2[group][sub_group][sub_group_2][Class]
                                    
                                    #Case of only one sub class (entry as string) or the signal word "all"
                                    if type(sub_classes)==str:                                                       
                                        if sub_classes==graph.nodes[node]['Sub_class'] or sub_classes=='all':    
                                            graph.add_node(node, sub_group_2=sub_group_2)
                                            grouped='Yes'
                                            break
                                            
                                    #Case of more than one sub classes (entry as list)        
                                    elif type(sub_classes)==list:
                                        for Subclass in sub_classes:
                                            if Subclass==graph.nodes[node]['Sub_class']:
                                                graph.add_node(node, sub_group_2=sub_group_2)
                                                grouped='Yes'
                                                break
                                        
            #To break out of all loops except the last to start the next sorting 
                                if grouped=='Yes':
                                    break                           
                            if grouped=='Yes':
                                break
                    if grouped=='Yes':
                        break
            if grouped=='Yes':
                break 
 
   
    #ASSIGN AGITATORS TO VESSELS

    Position={} 
    Agitators=[]
    
    #Create new attribute for every node
    for node in graph.nodes():
        if graph.nodes[node]['group']=='Vessel':
            graph.add_node(node, agitator='No')                
        else:
            graph.add_node(node, agitator='n.a.')
        
    #Select agitators and vessels        
    for node in graph.nodes():
        if graph.nodes[node]['Class']=='Agitator':
            Agitators.append(node)
            for other_node in graph.nodes():
                if graph.nodes[other_node]['group']=='Vessel':
                    
                    #Calculate distance by using coordinates
                    x_agitator=graph.nodes[node]['X']
                    x_other_node=graph.nodes[other_node]['X']
                    y_agitator=graph.nodes[node]['Y']
                    y_other_node=graph.nodes[other_node]['Y']
                    Position_difference=abs(x_agitator-x_other_node)+abs(y_agitator-y_other_node)#Calculate Position difference
                    Position[Position_difference]=other_node#save distance together with node
            
            #identificate nearest node to agitator and notice it in the attribute        
            node_assign=Position[min(Position.keys())]
            graph.add_node(node_assign, agitator='Yes')
    
    graph.remove_nodes_from(Agitators)
    
    
    #ASSIGN ORIFICE PLATE TO FLANGE
    
    Position={} 
    Orifice_plates=[]
 
    #Create new attribute for every node
    for node in graph.nodes():
        if graph.nodes[node]['Class']=='Flange':
            graph.add_node(node, orifice_plate='No')                
        else:
            graph.add_node(node, orifice_plate='n.a.')    
 
    #Select orifice plates and flanges        
    for node in graph.nodes():
        if graph.nodes[node]['Class']=='Orifice plate':
            Orifice_plates.append(node)
            for other_node in graph.nodes():
                if graph.nodes[other_node]['Class'] in ['Flange']:
                    
                    #calculate distance by using coordinates
                    x_flange=float(graph.nodes[node]['X'])
                    x_other_node=float(graph.nodes[other_node]['X'])
                    y_fange=float(graph.nodes[node]['Y'])
                    y_other_node=float(graph.nodes[other_node]['Y'])
                    Position_difference=abs(x_flange-x_other_node)+abs(y_fange-y_other_node)
                    Position[Position_difference]=other_node#save distance togehter with node
                           
            #identificate node with the smallest distance
            #reset the attributes
            #set the attributes of the orifice plate (Recent ID keeps the same until relabeling)
            node_assign=Position[min(Position.keys())]        
            graph.add_node(node_assign, orifice_plate='Yes')
    
    graph.remove_nodes_from(Orifice_plates)


    #REMOVE ISOLATED NODES
    
    nodes_isolated={} 
    
    #Identificate every isolated node which is not an Agitator or orifice plate      
    for node in graph.nodes():
        if nx.is_isolate(graph, node) and graph.nodes[node]['Class'] not in ['Agitator', 'Orifice plate']:
            nodes_isolated[node]=graph.nodes[node]['PID_label']
    
    #Showing a warning if necessary        
    if nodes_isolated!={}:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='Isolated nodes were identificated and removed. Please make sure it is correct.'
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_isolated)
        Counter_error+=1    
        
    graph.remove_nodes_from(nodes_isolated.keys())#remove isolated nodes from graph 
    
    
    #AVOID NO-LABEL NODES
     
    node_name=nx.get_node_attributes(graph, 'PID_label')
    no_label_nodes=[]
    double_names=[]
    
    #Identificate no-label nodes and give them the ID as name
    for node in graph.nodes():
        if node_name[node]=='n.d.':
            node_name[node]=node
            no_label_nodes.append(node)  
    
    #Warning
    if no_label_nodes!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='Warning: At least one node is not labeled. Label of such a node was exchanged with ID of DEXPI-Export. To avoid it, please add a label for the node in the P&ID.'   
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(no_label_nodes)
        Counter_error+=1
        
        
    #AVOID SAME NAME
    
    i=0
    
    #Identificate more time names
    for node_1 in graph.nodes():
        i+=1
        j=0
        for node_2 in graph.nodes():
            j+=1
            if node_name[node_1]==node_name[node_2] and i!=j and node_name[node_1] not in double_names:#avoid similarity because of same node and noticing a name for more than one time in the list
                double_names.append(node_name[node_1])
    
    #Identificate nodes with the more time names and give them consecutive numbers            
    for name in double_names:
        i=1           
        for node in graph.nodes():
            if node_name[node]==name:
                node_name[node]=name+' ('+str(i)+')'
                i+=1
    
    #Warning
    if double_names!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='At least one label is used more than one time. That is not possible cause a clear assignment must be given. For this run the nodes are numbered. For next run, please use clear label.'   
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(double_names)
        Counter_error+=1  
              
    
    #RELABELING
                  
    graph=nx.relabel_nodes(graph, node_name)
                     
                            
    #INVALID DRAWING (Example: Only inlets or outlets in valve)
    
    neighbors=[]
    all_neighbors=[]
    nodes_wrong_connection=[]
    
    #Notice in/out edges for every relevant nodes (except Vessel and MSR)
    for node in graph.nodes():
        if graph.nodes[node]['group']!='Vessel' and 'PIF' not in graph.nodes[node]['DEXPI_ID']:
            all_neighbors=list(nx.all_neighbors(graph, node))
            neighbors=list(nx.neighbors(graph, node))
            
            #every node with at least two edges and having only inlets or only outlets means invalid drawing and is noticed
            if len(all_neighbors)>1:#there must be more than one edge
                if len(neighbors)==0 or len(all_neighbors)-len(neighbors)==0:
                    nodes_wrong_connection.append(node)
    
    #Warning           
    if nodes_wrong_connection!=[]:
        Error_log.cell(row=Counter_error, column=Header_index['Warning/Error']).value='At least one node is wrong connected in the P&ID. Please check P&ID for the next run'
        Error_log.cell(row=Counter_error, column=Header_index['Equipment']).value=str(nodes_wrong_connection)
        Counter_error+=1
    
                       
    #ID-LIST
                
    for node in graph.nodes():
        ID_list.cell(row=Counter_ID_list, column=Header_index_ID_list['DEXPI_ID']).value=graph.nodes[node]['DEXPI_ID']
        ID_list.cell(row=Counter_ID_list, column=Header_index_ID_list['PID_label']).value=node
        ID_list.cell(row=Counter_ID_list, column=Header_index_ID_list['Neighbors']).value=str(list(nx.all_neighbors(graph, node)))[1:-1].replace("'", "")        
        ID_list.cell(row=Counter_ID_list, column=Header_index_ID_list['Descript']).value=graph.nodes[node]['descript'] 
        Counter_ID_list+=1         
                       
    #Save 
    nx.write_graphml(graph, Path_graph, encoding='utf-8', prettyprint=True, infer_numeric_types=False)#save the plot
    Template_workbook.save(Path_results)        
    return [graph]  

