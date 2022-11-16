# -*- coding: utf-8 -*-
"""
Created on Thu Dec  9 08:14:14 2021

@Author: Tim Holtermann, TU Dortmund, BCI AG Apparatedesign
"""

def eHazop(Path_graph, Path_HAZOP_data, Path_results):
    import networkx as nx
    import pandas as pd
    from Check_path import Check_path
    from openpyxl import load_workbook

    graph=nx.read_graphml(Path_graph)
    HAZOP_data=pd.read_excel(Path_HAZOP_data, index_col = 0)
    
    #WORKBOOKS    
    Results_workbook=load_workbook(Path_results)
    Results=Results_workbook['Results'] 
    Error_log=Results_workbook['Error_log']
    
    max_col = Results.max_column
    Header_index_results={}  
    for i in range(1, max_col + 1): 
        Header = Results.cell(row = 1, column = i).value
        Header_index_results[Header]=i
    for row in range(1, 100):        
        if Results.cell(row=row, column=Header_index_results['Index']).value==None:        
            Counter_results=row
            break 
        
    max_col = Error_log.max_column
    Header_index_error={}  
    for i in range(1, max_col + 1): 
        Header = Error_log.cell(row = 1, column = i).value
        Header_index_error[Header]=i    
    for row in range(1, 100):        
        if Error_log.cell(row=row, column=Header_index_error['Warning/Error']).value==None:        
            Counter_error=row
            break
        
    
    #CHECK DESIGN PARAMETER BEFORE HAZOP
       
    Missing_data_nodes=[]                             
    for node in graph.nodes:
        if graph.nodes[node]['group'] in ['Vessel', 'Column', 'Filter', 'Pump', 'Heat exchanger', '...']:
            if graph.nodes[node]['P_max_operation'] not in ['n.a.', 'n.d.'] and graph.nodes[node]['P_max_design'] not in ['n.a.', 'n.d.']: 
                if graph.nodes[node]['P_max_operation']>graph.nodes[node]['P_max_design']:                    
                    Error_log.cell(row=Counter_error, column=Header_index_error['Warning/Error']).value='Max Design Pressure is not sufficient.'      
                    Error_log.cell(row=Counter_error, column=Header_index_error['Equipment']).value=node
                    Counter_error+=1
            else:
                Missing_data_nodes.append(node)
                
            if graph.nodes[node]['P_min_operation'] not in ['n.a.', 'n.d.'] and graph.nodes[node]['P_min_design'] not in ['n.a.', 'n.d.']: 
                if graph.nodes[node]['P_min_operation']<graph.nodes[node]['P_min_design']:
                    Error_log.cell(row=Counter_error, column=Header_index_error['Warning/Error']).value='Min Design Pressure is not sufficient.'
                    Error_log.cell(row=Counter_error, column=Header_index_error['Equipment']).value=node                        
                    Counter_error+=1
            else:
                Missing_data_nodes.append(node)
            if graph.nodes[node]['T_max_operation'] not in ['n.a.', 'n.d.'] and graph.nodes[node]['T_max_design'] not in ['n.a.', 'n.d.']: 
                if graph.nodes[node]['T_max_operation']>graph.nodes[node]['T_max_design']:
                    Error_log.cell(row=Counter_error, column=Header_index_error['Warning/Error']).value='Max Design Temperature is not sufficient.'      
                    Error_log.cell(row=Counter_error, column=Header_index_error['Equipment']).value=node
                    Counter_error+=1
            else:
                Missing_data_nodes.append(node)
            if graph.nodes[node]['T_min_operation'] not in ['n.a.', 'n.d.'] and graph.nodes[node]['T_min_design'] not in ['n.a.', 'n.d.']: 
                if graph.nodes[node]['T_min_operation']<graph.nodes[node]['T_min_design']:
                    Error_log.cell(row=Counter_error, column=Header_index_error['Warning/Error']).value='Min Design Temperature is not sufficient.'      
                    Error_log.cell(row=Counter_error, column=Header_index_error['Equipment']).value=node
                    Counter_error+=1
            else:
                Missing_data_nodes.append(node)

    if Missing_data_nodes!=[]:
        Error_log.cell(row=Counter_error, column=Header_index_error['Warning/Error']).value='No Process Parameter Checking possible. Simulation data or Equipment data are missing.'      
        Error_log.cell(row=Counter_error, column=Header_index_error['Equipment']).value=str(set(Missing_data_nodes))[1:-1]               
        Counter_error+=1
                 
            
    #READING DATA STORAGE (row after row)
    
    for index, row in HAZOP_data.iterrows():
        
        #read the one-time entries
        Deviation=HAZOP_data.loc[index, 'Guideword']+' '+HAZOP_data.loc[index, 'Parameter']
        Description=HAZOP_data.loc[index, 'Description']
        Cause=HAZOP_data.loc[index, 'Cause']
        Consequence=HAZOP_data.loc[index, 'Consequence']          
        Propability=HAZOP_data.loc[index, 'Propability']         
        Requirements=HAZOP_data.loc[index, 'Requirements']
        Safeguard_1=HAZOP_data.loc[index, 'Safeguard_1']
        Safeguard_2=HAZOP_data.loc[index, 'Safeguard_2']        

                                
        #read the mutiple entries, one after one
        Equipments=[]
        Equipments_specifications=[] 
        #Connections_specifications=[] 
        for j in range(1,4):#max 3 equipments
        
            #Reading the entries only if the equipment entry is not empty
            if type(HAZOP_data.loc[index, 'Equipment_'+str(j)])==str:
                
                Equipment=HAZOP_data.loc[index, 'Equipment_'+str(j)]
                Equipments.append(Equipment)#add equipment etnry as a list
                
                Equipment_specifications=[]
                Specification_1=HAZOP_data.loc[index, 'Specification_'+str(j)+'.1']
                Specification_2=HAZOP_data.loc[index, 'Specification_'+str(j)+'.2']   
                if type(Specification_1)==str:
                    Equipment_specifications.append(Specification_1)
                if type(Specification_2)==str:                
                    Equipment_specifications.append(Specification_2)
                Equipments_specifications.append(Equipment_specifications)
        
        
        #CONVERT EQUIPMENT IN NODES
        
        Equipments_nodes=[]        
        for j in range(0, len(Equipments)):#for all equipment entries
            nodes_possible=[]#to collect the nodes
            
            #Find nodes with the right (sub) group
            for node in graph.nodes():                
                if graph.nodes[node]['group'] == Equipments[j] or graph.nodes[node]['sub_group'] == Equipments[j] or graph.nodes[node]['sub_group_2'] == Equipments[j] or 'All' == Equipments[j]:#all equipments with the right (sub) group or all if all                
                    
                    #if ther are no specifications, the nodes can be collected immediately   
                    if Equipments_specifications[j]==[]:                    
                        nodes_possible.append(node)        
                        
                    #Otherwise check the specifications    
                    else:
                        Equipments_specifications_check=Equipments_specifications[j].copy()         
                        if 'Jacket' in Equipments_specifications_check and graph.nodes[node]['C_H_system'] in ['full-tube heating cooling coil', 'heating-cooling jacket', 'Semi-tube heating cooling coil']:                   
                            Equipments_specifications_check.remove('Jacket')
                        if 'Electric heat system' in Equipments_specifications_check and graph.nodes[node]['C_H_system'] in ['Electrical heating']:                   
                            Equipments_specifications_check.remove('Electric heating')                              
                        if 'Inert' in Equipments_specifications_check and graph.nodes[node]['inert']=='Yes':                  
                            Equipments_specifications_check.remove('Inert')                          
                        if 'Rotating part' in Equipments_specifications_check and graph.nodes[node]['rotating_part']=='Yes':                    
                            Equipments_specifications_check.remove('Rotating part')
                        if 'Agitator' in Equipments_specifications_check and graph.nodes[node]['agitator']=='Yes':                    
                            Equipments_specifications_check.remove('agitator')                            
                        if 'Level Control' in Equipments_specifications_check:
                            if 'Level Control' in graph.nodes[node]['measurements']:                    
                                Equipments_specifications_check.remove('Level Control')
                            elif 'Level Control' in graph.nodes[node]['signals']:                    
                                Equipments_specifications_check.remove('Level Control')
                        
                        '''        
                        if 'Flow Control' in Equipments_specifications_check and 'Flow Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Flow Control')                                                
                        if 'Temperature Control' in Equipments_specifications_check and 'Temperature Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Temperature Control')                                              
                        if 'Pressure Control' in Equipments_specifications_check and 'Pressure Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Pressure Control')
                        if 'Inlet Flow Control' in Equipments_specifications_check and 'Inlet Flow Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Inlet Flow Control')
                        if 'Outlet Flow Control' in Equipments_specifications_check and 'Outlet Flow Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Outlet Flow Control')                                                
                        if 'Inlet Temperature Control' in Equipments_specifications_check and 'Inlet Temperature Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Inlet Temperature Control')
                        if 'Outlet Temperature Control' in Equipments_specifications_check and 'Outlet Temperature Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Outlet Temperature Control')
                        if 'Inlet Pressure Control' in Equipments_specifications_check and 'Inlet Pressure Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Inlet Pressure Control')                                                
                        if 'Outlet Pressure Control' in Equipments_specifications_check and 'Outlet Pressure Control' in graph.nodes[node]['controls']:                    
                            Equipments_specifications_check.remove('Outlet Pressure Control')
                            '''                            
                        #an empty list means all specifications are given, so the node can be cpllected
                        if  len(Equipments_specifications_check)==0:
                                nodes_possible.append(node)
                                
            #Add collected nodes            
            Equipments_nodes.append(nodes_possible)
            
            #check if not all necessary equipments were found
            if nodes_possible==[]:#if any Equipment was not found, at least one entry in the Equipments_nodes is empty
                exist='No'
                print('Equipment '+str(j)+' not given (index '+str(index)+')')
                break
            else:
                exist='Yes'
                       
        if exist!='No':
            print(Equipments_nodes)#for comprehension (end situation)           
            if len(Equipments)==1:#non-path cases
                for node in Equipments_nodes[0]:
                    
                    Results.cell(row=Counter_results, column=Header_index_results['Index']).value=index 
                    Results.cell(row=Counter_results, column=Header_index_results['Description']).value=Description                      
                    Results.cell(row=Counter_results, column=Header_index_results['Deviation']).value=Deviation                                                             
                    Results.cell(row=Counter_results, column=Header_index_results['Cause']).value=Cause
                    Results.cell(row=Counter_results, column=Header_index_results['Consequence']).value=Consequence
                    Results.cell(row=Counter_results, column=Header_index_results['Involved Equipment']).value=graph.nodes[node]['PID_label']                    
                    Results.cell(row=Counter_results, column=Header_index_results['F']).value=Propability                        
                    Results.cell(row=Counter_results, column=Header_index_results['Requirements']).value=Requirements
                    Counter_results+=1       
                    
            elif len(Equipments)>1:#Path have to be found
                       
                #select first and second equipment nodes and finding paths between them
                for node_1 in Equipments_nodes[0]: 
                    for node_2 in Equipments_nodes[1]:
                        nodes_regarded=[node_1, node_2]
                        nodes_between_forbidden=Equipments_nodes[0]+Equipments_nodes[1]#Total list of the nodes, which must not appear in between the paths (because those nodes must be only start/end of paths)
                        Hazop_case='Yes'
                        paths=list(nx.all_simple_paths(graph, node_1, node_2))#all possible paths 
                        
                        #proving one path after the other
                        for path in paths:
                            Return=Check_path(graph, path, 0, nodes_between_forbidden, "", "", index, 'No')
                            path_okay=Return[0]
                            path_class=Return[1]
                            path_sub_class=Return[2]
                            path_change=Return[3]                            
                            if path_okay=='No':
                                Hazop_case='No'
                        
                            #In case of validity and a two-equipment case, HAZOP case is added                       
                            if len(Equipments)==2 and Hazop_case=='Yes': 
                                Results.cell(row=Counter_results, column=Header_index_results['Index']).value=index
                                Results.cell(row=Counter_results, column=Header_index_results['Involved Equipment']).value=graph.nodes[node_1]['PID_label']+', '+graph.nodes[node_2]['PID_label']
                                Results.cell(row=Counter_results, column=Header_index_results['Deviation']).value=Deviation  
                                Results.cell(row=Counter_results, column=Header_index_results['Description']).value=Description                                                                                 
                                Results.cell(row=Counter_results, column=Header_index_results['Cause']).value=Cause
                                Results.cell(row=Counter_results, column=Header_index_results['Consequence']).value=Consequence
                                Results.cell(row=Counter_results, column=Header_index_results['F']).value=Propability
                                Results.cell(row=Counter_results, column=Header_index_results['Requirements']).value=Requirements
                                Counter_results+=1
    
                            #otherwise a correct path to he next equipment has to be found          
                            elif len(Equipments)>2:
                                nodes_between_forbidden=Equipments_nodes[1]+Equipments_nodes[2]#like above                         
                                
                                #find the possible paths to the next equipment
                                for node_3 in Equipments_nodes[2]:#every associated node
                                    del nodes_regarded[2:]
                                    nodes_regarded.append(node_3)
                                    paths_2=nx.all_simple_paths(graph, node_2, node_3)
                                    
                                    #check the paths
                                    for path_2 in paths_2:
                                        Return=Check_path(graph, path_2, 1, nodes_between_forbidden, path_class, path_sub_class, index, path_change)
                                        path_okay=Return[0]
                                        path_class=Return[1]
                                        path_sub_class=Return[2] 
                                        path_change=Return[3]  
                                        if path_okay=='No':
                                            Hazop_case='No'  
                                        
     
                                        #In case of validity and a three-equipment case, HAZOP case is added 
                                        if Hazop_case=='Yes':
                                            print('selected: '+str(path+path_2[1:]))#[1:] to avoid a double node (end and start of the different paths are equal)

                                            Results.cell(row=Counter_results, column=Header_index_results['Index']).value=index
                                            Results.cell(row=Counter_results, column=Header_index_results['Involved Equipment']).value=graph.nodes[node_1]['PID_label']+', '+graph.nodes[node_2]['PID_label']+', '+graph.nodes[node_3]['PID_label']
                                            Results.cell(row=Counter_results, column=Header_index_results['Deviation']).value=Deviation  
                                            Results.cell(row=Counter_results, column=Header_index_results['Description']).value=Description                                                                                 
                                            Results.cell(row=Counter_results, column=Header_index_results['Cause']).value=Cause
                                            Results.cell(row=Counter_results, column=Header_index_results['Consequence']).value=Consequence
                                            Results.cell(row=Counter_results, column=Header_index_results['F']).value=Propability
                                            Results.cell(row=Counter_results, column=Header_index_results['Requirements']).value=Requirements
                                            Counter_results+=1 
                                                  
                                                                          
    Results_workbook.save(Path_results)

    #SAFEGUARDS
   
    #Read the sheets from Results    
    Results_workbook=load_workbook(Path_results)
    Results=Results_workbook['Results']
    Error_log=Results_workbook['Error_log']
    
    #Assign index to header (results)
    max_col=Results.max_column
    Header_index_results={}  
    for i in range(1, max_col + 1): 
        Header = Results.cell(row = 1, column = i).value
        Header_index_results[Header]=i  
    
    #Assign index to header and check in which row the list ends (error)
    max_col = Error_log.max_column
    Header_index_error={}  
    for i in range(1, max_col + 1): 
        Header = Error_log.cell(row = 1, column = i).value
        Header_index_error[Header]=i    
    for row in range(1, 100):        
        if Error_log.cell(row=row, column=Header_index_error['Warning/Error']).value==None:        
            Counter_error=row
            break  
        
        
    #Going through result scenarios
    row=2
    while type(Results.cell(row=row, column=Header_index_results['Description']).value)==str:
        
        #All relevant entries in results
        Index=Results.cell(row=row, column=Header_index_results['Index'])
        Equipment=Results.cell(row=row, column=Header_index_results['Involved Equipment'])
        Safeguards_all=Results.cell(row=row, column=Header_index_results['Potencial Safeguards'])
        Safeguards_useful=Results.cell(row=row, column=Header_index_results['Existing Safeguards'])
        Safeguards_suggested=Results.cell(row=row, column=Header_index_results['Suggested Safeguards'])
        
        #Relevant entries from scenario data base        
        Safeguards=[]
        Safeguard_1=HAZOP_data.loc[Index.value, 'Safeguard_1']
        Safeguard_2=HAZOP_data.loc[Index.value, 'Safeguard_2']        
        if type(Safeguard_1)==str:            
            Safeguards.append(Safeguard_1)
            Safeguards_suggested.value=Safeguard_1
        if type(Safeguard_2)==str:            
            Safeguards.append(Safeguard_2)
            Safeguards_suggested.value=Safeguards_suggested.value+', '+Safeguard_2
                            
        #Print all existing safeguards of nodes mentioned in the scenario
        Existing=[]        
        #All mentioned equipments which has safeguards
        for node in Equipment.value.split(', '):
            if graph.nodes[node]['safeguards'] not in ['n.a.', 'No']:
                
                #Print the list of all safeguards (for every mentioned node)
                if Safeguards_all.value==None:
                    Safeguards_all.value=node+': '+graph.nodes[node]['safeguards']
                else:
                    Safeguards_all.value=Safeguards_all.value+'// '+node+': '+graph.nodes[node]['safeguards']  
                
                #In case of suggested Safeguards from HAZOP data
                #If there is a useful safefuard in All_safeguards print it
                if Safeguards!=[]:
                    
                    #compare suggested with all existing
                    for Safeguard_1 in graph.nodes[node]['safeguards'].split(', '):
                        for Safeguard_2 in Safeguards:
                            if Safeguard_2 in Safeguard_1:
                                Existing.append(Safeguard_2)
                                
                                #Entry
                                if Safeguards_useful.value==None:
                                    Safeguards_useful.value=Safeguard_1+' ('+node+')'
                                else:            
                                    Safeguards_useful.value=Safeguards_useful.value+', '+Safeguard_1+' ('+node+')'
                                                            
        row+=1   

             
    Results_workbook.save(Path_results)    