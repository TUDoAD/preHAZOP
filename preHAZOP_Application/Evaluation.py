# -*- coding: utf-8 -*-
"""
Created on Mon Jan 10 09:10:58 2022

@Author: Tim Holtermann, TU Dortmund, BCI AG Apparatedesign
"""

def Risk_assesement(Path_graph, Path_results, Path_HAZOP_data):
    import networkx as nx
    import pandas as pd
    from openpyxl import load_workbook
    
    Graph=nx.read_graphml(Path_graph)   
    
    #Read the sheets from data storage
    HAZOP_data=pd.read_excel(Path_HAZOP_data, index_col = 0)    
    Substance_data=pd.read_excel(Path_HAZOP_data, sheet_name='Substance', index_col = 0)
    Equipment_costs=pd.read_excel(Path_HAZOP_data,sheet_name='Equipment', index_col = 0)
    Risk_matrix=pd.read_excel(Path_HAZOP_data,sheet_name='Risk Matrix', index_col = 0)
    Hazard_leakage=pd.read_excel(Path_HAZOP_data,sheet_name='Leakage Matrix', index_col = 0)  
    Hazard_volume=pd.read_excel(Path_HAZOP_data,sheet_name='Volume Matrix', index_col = 0) 
    Finance_severity=pd.read_excel(Path_HAZOP_data,sheet_name='Finance', index_col = 0)  
    Data_workbook=load_workbook(Path_HAZOP_data)
        
    #Read the sheets from Results    
    Results_workbook=load_workbook(Path_results)
    Results=Results_workbook['Results']
    Error_log=Results_workbook['Error_log']
    
    #Assign index to header (results)
    max_col=Results.max_column
    Header_index_results={}  
    for i in range(1, max_col + 1): 
        Header = Results.cell(row = 1, column = i).value
        Header_index_results[Header]=i  
    
    #Assign index to header and check in which row the list ends (error)
    max_col = Error_log.max_column
    Header_index_error={}  
    for i in range(1, max_col + 1): 
        Header = Error_log.cell(row = 1, column = i).value
        Header_index_error[Header]=i    
    for row in range(1, 100):        
        if Error_log.cell(row=row, column=Header_index_error['Warning/Error']).value==None:        
            Counter_error=row
            break  
        
            
    #ITERATE SCENARIO RESULTS
    row=2
    while type(Results.cell(row=row, column=Header_index_results['Description']).value)==str:
        
        #All relevant entries in results
        Index=Results.cell(row=row, column=Header_index_results['Index'])
        Propability_1=Results.cell(row=row, column=Header_index_results['F'])
        Severity_1=Results.cell(row=row, column=Header_index_results['S'])
        Risk_1=Results.cell(row=row, column=Header_index_results['Risk'])
        Equipment=Results.cell(row=row, column=Header_index_results['Involved Equipment'])
        Cause=Results.cell(row=row, column=Header_index_results['Cause'])
        Description=Results.cell(row=row, column=Header_index_results['Description'])
        Consequences=Results.cell(row=row, column=Header_index_results['Consequence'])
        Substance=Results.cell(row=row, column=Header_index_results['Substance'])
        
        #Relevant entries from scenarios          
        Equipment_affected=HAZOP_data.loc[Index.value, 'Affected Equipment']
                             
        #predefine severity from categories
        S_loss='n.d.'
        S_volume='n.d.'
        S_leakage='n.d.'
        
        #Identificate affected equipment
        Equipment_number={'Equipment_1':0, 'Equipment_2':1, 'Equipment_3':2}
        Number=Equipment_number[Equipment_affected]
        Nodes_mentioned=Equipment.value.split(', ')
        Node_affected=Nodes_mentioned[Number]
        
    
        #SUBSTANCE INFORMATION 
        for Consequence in ['Rupture', 'Leakage', 'Seal leakage', 'Big Leakage']:             
            if Consequence in Consequences.value:
                
                #Finding Substances + hazards
                Subs=[]
                Subs=Graph.nodes[Node_affected]['Subs_process'].split(', ')
                Signal_words=[]
                for Sub in Subs:
                    Signal_words.append(Substance_data.loc[Sub, 'Signal word'])
                    hazards=[]
                    for GHS in ['explosive', 'flammable', 'oxidizing', 'compressed gas', 'corrosive', 'toxic','harmful','health hazard','environmental hazard']:
                        if Substance_data.loc[Sub, GHS]:
                            hazards.append(GHS)
                    
                    #Printing Substances + hazards
                    hazards=str(hazards)[1:-1].replace("'", "")                            
                    if type(Substance.value)!=str:
                        Substance.value=Sub+' ('+hazards+')'                        
                    else:                    
                        Substance.value=Substance.value+', '+Sub+' ('+hazards+')'
                        
                #Identificate most dangerous signal word        
                if 'Danger' in Signal_words:
                    Signal_word='Danger'
                elif 'Warning' in Signal_words:
                    Signal_word='Warning'
                elif '-' in Signal_words:
                    Signal_word='-'                     
                                            
                break


        #SEVERITY HAZARD/VOLUME
        for Consequence in ['Rupture', 'Leakage', 'Seal leakage', 'Big Leakage']:             
            if Consequence in Consequences.value: 
                
                #We need the quantitiy as a product of Volume and Density
                Flow=Graph.nodes[Node_affected]['Mass_flow']
                t=1200
                if Flow not in ['n.d.', 'n.a.']:                            
                    Quantity=float(Flow)*t
                    
                    for Volume_range in Hazard_volume.columns:
                        if Quantity<=Volume_range:
                            S_volume=Hazard_volume.loc[Signal_word, Volume_range]
                            break
                
                break
            
 
        #SEVERITY HAZARD/LEAKAGE
        for Consequence in ['Leakage', 'Seal leakage', 'Rupture']:             
            if Consequence == Consequences.value:
                S_leakage=Hazard_leakage.loc[Signal_word, Consequence]
                break
        
        
        #EQUIPMENT LOSS
        Loss_equipment=0
        for Consequence in ['Damage', '...']:            
            if Consequence in Consequences.value:
                if Graph.nodes[Node_affected]['group'] in ['Valves', 'Fittings', 'Pipe tee', '...']:
                    Loss_equipment=1#to be over zero for severity determination                    
                else:
                    try:
                        Loss_equipment=Equipment_costs.loc[Node_affected, 'Cost']
                    except KeyError:
                        print('The affected node '+Node_affected+' is not in the equipment list.')
                #Loss_equipment=input('Loss for Scenario '+Index.value+' ('+Description.value+') with the consequence(s) '+Consequence.value)
               
        
        #SEVERITY LOSS
        #Make sure the biggest loss is concidered
        for Consequence in ['Damage', '...']:            
            if Consequence in Consequences.value:
                Loss_before=0
                for Loss in [Loss_equipment]:
                    if Loss>Loss_before:
                        Loss_before=Loss
                        
                        for Loss_range in Finance_severity.index:
                            if Loss<=Loss_range:
                                S_loss=Finance_severity.loc[Loss_range, 'Severity']
                                break
                    
                    
        #MAX SEVERITY            
        S_number={'S0':0, 'S1':1, 'S2':2, 'S3':3, 'S4':4}
        Number_S= {0:'S0', 1:'S1', 2:'S2', 3:'S3', 4:'S4'}
        S_list=[]
        for S in [S_loss, S_volume, S_leakage]:
            if S!='n.d.':
                S_list.append(S_number[S])
        if S_list!=[]:
            Severity_1.value=Number_S[min(S_list)]    
            

        #RESULT  
        
        if Severity_1.value==None:
            print('No severity estimation could be made. Unknown consequence might be the reason.')
            #Severity_1.value=input('Severity for "'+Consequences.value+'" in Scenario "'+str(Index.value)+'" in row '+str(row)+'. Please type in the severity: ')
        elif Propability_1.value==None:
            print('There is no propability assumed for "'+Cause.value+'" in Scenario "'+Description.value+'" in row '+str(row))            
        else:            
            try:
                Risk_1.value=Risk_matrix.loc[Propability_1.value, Severity_1.value]
            except KeyError:
                print('Wrong input for severity or propability of scenario '+str(Index.value))
                              
        row+=1   

             
    Results_workbook.save(Path_results)
    
    